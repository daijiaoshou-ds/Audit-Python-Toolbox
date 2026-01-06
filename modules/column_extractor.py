"""
Excel数据提取器 - 升级版
修复了Excel读取兼容性bug，集成Polars高性能引擎
支持pandas和polars双引擎模式
fastexcel集成：真正发挥Polars的Rust级别性能
"""
import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading
from difflib import SequenceMatcher
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional, List, Dict, Any, Tuple, Set

# ==================== 新增：xlsxwriter支持检测 ====================
try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False
    print("提示: xlsxwriter未安装，写入速度可能较慢。可通过 'pip install xlsxwriter' 提升保存性能。")
try:
    import fastexcel
    FASTEXCEL_AVAILABLE = True
except ImportError:
    FASTEXCEL_AVAILABLE = False
    print("提示: fastexcel未安装。可通过 'pip install fastexcel' 安装以提升Polars性能。")

# ==================== 新增：Polars支持 ====================
try:
    import polars as pl
    POLARS_AVAILABLE = True
except ImportError:
    POLARS_AVAILABLE = False
    print("警告: Polars未安装，将使用pandas模式。可通过 'pip install polars' 安装以提升性能。")

# ==================== 新增：pyarrow支持检测 ====================
try:
    import pyarrow
    PYARROW_AVAILABLE = True
except ImportError:
    PYARROW_AVAILABLE = False
    print("警告: pyarrow未安装，Polars模式性能可能受限。可通过 'pip install pyarrow' 提升性能。")

# ==================== 核心逻辑层 ====================

def calculate_similarity(s1: str, s2: str) -> float:
    """计算两个字符串的相似度"""
    if not s1 or not s2:
        return 0.0
    return SequenceMatcher(None, s1.lower(), s2.lower()).ratio()

def is_fuzzy_match(target_keyword: str, cell_value: Any, threshold: float) -> bool:
    """智能模糊匹配判断"""
    t = str(target_keyword).strip().lower()
    v = str(cell_value).strip().lower() if cell_value is not None else ""
    if threshold >= 0.99:
        return t == v
    if threshold <= 0.9:
        if t in v or v in t:
            return True
    score = calculate_similarity(t, v)
    return score >= threshold

def get_files_to_process(source_path: str, is_folder: bool, recursive: bool) -> List[str]:
    """根据设置获取文件列表"""
    files_list = []
    if not is_folder:
        if os.path.isfile(source_path) and source_path.lower().endswith(('.xlsx', '.xlsm')):
            return [source_path]
        return []
    if not os.path.exists(source_path):
        return []

    if recursive:
        for root, dirs, files in os.walk(source_path):
            for f in files:
                if f.lower().endswith(('.xlsx', '.xlsm')) and not f.startswith("~$"):
                    files_list.append(os.path.join(root, f))
    else:
        for f in os.listdir(source_path):
            full_path = os.path.join(source_path, f)
            if os.path.isfile(full_path) and f.lower().endswith(('.xlsx', '.xlsm')) and not f.startswith("~$"):
                files_list.append(full_path)
    return files_list

# ==================== 【修复】新增：Excel文件预处理 ====================
def repair_excel_file(file_path: str, log_func) -> bool:
    """
    修复损坏或格式不标准的Excel文件
    通过重新保存来修复文件的元数据问题
    """
    try:
        # 首先尝试正常读取
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # 检查维度是否合理
        if ws.max_row > 0 and ws.max_column > 0:
            # 尝试读取第一行数据来验证文件可读性
            test_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if test_row and any(cell is not None for cell in test_row[0] if test_row):
                wb.close()
                return True  # 文件正常
        wb.close()
    except Exception as e:
        log_func(f"检测到文件异常，尝试修复: {os.path.basename(file_path)}")
    
    try:
        # 尝试修复：使用非只读模式打开并重新保存
        temp_wb = load_workbook(file_path, read_only=False, data_only=True)
        temp_file = file_path + ".temp"
        
        # 保存到临时文件
        temp_wb.save(temp_file)
        temp_wb.close()
        
        # 替换原文件
        shutil.move(temp_file, file_path)
        log_func(f"✓ 文件修复成功: {os.path.basename(file_path)}")
        return True
    except Exception as e:
        log_func(f"✗ 文件修复失败: {os.path.basename(file_path)} - {str(e)}")
        return False

def scan_header_and_map_columns(worksheet, mode: str, 
                                exact_cols: Optional[List[str]] = None, 
                                fuzzy_cols: Optional[List[str]] = None, 
                                fuzzy_threshold: float = 0.6,
                                scan_rows: int = 15) -> Dict[int, str]:
    """扫描表头并建立映射"""
    max_scan = min(worksheet.max_row, scan_rows)
    if max_scan == 0:
        return {}
    best_mapping = {}
    max_matches = -1
    
    for r in range(1, max_scan + 1):
        row_values = []
        for c in range(1, worksheet.max_column + 1):
            val = worksheet.cell(row=r, column=c).value
            row_values.append(str(val).strip() if val is not None else "")

        if all(v == "" for v in row_values):
            continue
        current_mapping = {}
        
        if mode == 'all':
            name_counter = {}
            for c_idx, val in enumerate(row_values, 1):
                if not val:
                    continue
                count = name_counter.get(val, 0)
                final_name = val if count == 0 else f"{val}_{count}"
                name_counter[val] = count + 1
                current_mapping[c_idx] = final_name
            if current_mapping:
                return current_mapping

        else:
            matches_count = 0
            name_counter = {}
            exact_cols = exact_cols or []
            fuzzy_cols = fuzzy_cols or []
            
            for c_idx, val in enumerate(row_values, 1):
                if not val:
                    continue
                val_lower = val.lower()
                matched_name = None
                
                # 精确匹配
                for target in exact_cols:
                    if target.lower() == val_lower:
                        matched_name = target
                        break
                
                # 模糊匹配
                if not matched_name and fuzzy_cols:
                    for target in fuzzy_cols:
                        if is_fuzzy_match(target, val, fuzzy_threshold):
                            matched_name = target
                            break
                
                if matched_name:
                    count = name_counter.get(matched_name, 0)
                    final_key = matched_name if count == 0 else f"{matched_name}_{count}"
                    name_counter[matched_name] = count + 1
                    current_mapping[c_idx] = final_key
                    matches_count += 1
            
            if matches_count > max_matches:
                max_matches = matches_count
                best_mapping = current_mapping
    
    return best_mapping

def read_excel_with_polars(file_path: str, fname: str, log_func=None) -> Tuple[List[pl.DataFrame], Set[str], int, float]:
    """
    使用Polars+fastexcel引擎极速读取单个Excel文件 - Arrow零拷贝版本
    
    优化策略：
    1. 直接使用Polars读取（自动使用fastexcel引擎）
    2. 先快速获取所有sheet名称（openpyxl只读模式很快）
    3. 然后用fastexcel逐个读取所有sheet
    4. 使用Polars DataFrame原生格式，避免转换为Python字典（零拷贝）
    5. 使用with_columns添加来源信息（高效操作）
    
    Returns: (DataFrame列表, 所有列名集合, 处理的工作表数量, 读取耗时)
    """
    t_start = time.time()
    
    # 快速获取所有工作表名称（openpyxl只读模式很快）
    sheet_names = ["Sheet1"]  # 默认
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = [ws.title for ws in wb.worksheets]
        wb.close()
    except:
        pass
    
    all_dfs = []
    all_columns = set()
    sheet_count = 0
    
    for sheet_name in sheet_names:
        try:
            # 使用Polars读取（fastexcel引擎，自动选择最优方式）
            df = pl.read_excel(file_path, sheet_name=sheet_name)
            
            if df is not None and df.height > 0:
                # 使用Polars原生方式添加来源列（零拷贝操作）
                df = df.with_columns([
                    pl.lit(fname).alias('_来源工作簿'),
                    pl.lit(sheet_name).alias('_来源工作表')
                ])
                
                all_dfs.append(df)
                all_columns.update(df.columns)
                sheet_count += 1
                
        except Exception as e:
            # 读取失败，尝试指定fastexcel引擎
            try:
                df = pl.read_excel(file_path, sheet_name=sheet_name, engine="fastexcel")
                if df is not None and df.height > 0:
                    df = df.with_columns([
                        pl.lit(fname).alias('_来源工作簿'),
                        pl.lit(sheet_name).alias('_来源工作表')
                    ])
                    all_dfs.append(df)
                    all_columns.update(df.columns)
                    sheet_count += 1
            except:
                pass
    
    elapsed = time.time() - t_start
    
    if log_func and all_dfs:
        total_rows = sum(df.height for df in all_dfs)
        log_func(f"  ✓ 读取完成: 共 {total_rows} 行，来自 {sheet_count} 个工作表 (耗时 {elapsed:.2f}s)")
    elif log_func and not all_dfs:
        log_func(f"  ✗ 读取失败，无数据")
    
    return all_dfs, all_columns, sheet_count, elapsed


def read_excel_parallel(file_path: str, fname: str, log_func=None):
    """
    并行读取的辅助函数 - 返回结构化结果
    
    用于concurrent.futures并行处理
    """
    dfs, columns, count, elapsed = read_excel_with_polars(file_path, fname, log_func)
    return {
        'file_path': file_path,
        'fname': fname,
        'dfs': dfs,
        'columns': columns,
        'count': count,
        'elapsed': elapsed,
        'rows': sum(df.height for df in dfs) if dfs else 0
    }


def normalize_dataframes_for_concat(dfs: List[pl.DataFrame], target_columns: List[str], log_func) -> List[pl.DataFrame]:
    """
    规范化DataFrame列表，使它们具有相同的schema（列结构）
    
    这是驯服Arrow的关键！
    - 确保每个DataFrame都有目标列
    - 缺失的列填充空值（None）
    - 保持Arrow格式不转换为Python对象
    
    Args:
        dfs: DataFrame列表
        target_columns: 目标列列表（所有文件共有的列 + 来源列）
        log_func: 日志函数
    
    Returns:
        规范化后的DataFrame列表，可以安全使用pl.concat()合并
    """
    if not dfs:
        return dfs
    
    log_func(f"  🔧 正在规范化 {len(dfs)} 个DataFrame的schema...")
    
    normalized_dfs = []
    total_added_cols = 0
    
    for i, df in enumerate(dfs):
        df_cols = set(df.columns)
        target_set = set(target_columns)
        
        # 找出缺失的列
        missing_cols = target_set - df_cols
        
        if missing_cols:
            # 添加缺失的列，填充None（空值）
            # 使用with_columns可以高效地添加列（Arrow操作）
            add_exprs = []
            for col in missing_cols:
                add_exprs.append(pl.lit(None).alias(col))
                total_added_cols += 1
            
            df = df.with_columns(add_exprs)
        
        # 确保列顺序一致
        df = df.select(target_columns)
        normalized_dfs.append(df)
    
    if total_added_cols > 0:
        log_func(f"  ✓ 规范化完成: 共补全 {total_added_cols} 个缺失列")
    
    return normalized_dfs

def process_with_polars(files: List[str], target_sheets: List[str], 
                       extract_mode: str, exact_cols: List[str], 
                       fuzzy_cols: List[str], fuzzy_threshold: float,
                       save_path: str, log_func, stop_event=None,
                       is_csv: bool = False) -> Tuple[bool, str]:
    """
    使用Polars引擎处理数据 - fastexcel真正高性能版本
    支持CSV极速保存模式
    
    性能优势：
    - fastexcel基于Rust calamine，比openpyxl快5-10倍
    - CSV保存比Excel快5-10倍
    - 直接读取到Apache Arrow格式，零拷贝转换
    - 内存占用约为openpyxl的1/10
    - fastexcel内部已实现Rust层面的多线程优化
    """
    
    log_func("=== Polars高性能模式 (fastexcel引擎) ===")
    
    # 显示引擎信息
    if FASTEXCEL_AVAILABLE:
        log_func("🚀 使用fastexcel引擎（Rust calamine）- 真正的极速体验")
    else:
        log_func("⚠️ fastexcel未安装，将使用openpyxl引擎")
        log_func("💡 安装命令: pip install fastexcel")
    
    log_func(f"发现 {len(files)} 个文件...")
    
    # 🚀 智能并行读取配置
    # 策略：fastexcel内部已对大文件多线程优化，
    # 外部并行只对中小文件有效，大文件顺序读避免线程开销
    
    # 统计文件信息
    file_info = []
    for f in files:
        size = os.path.getsize(f)
        file_info.append({'path': f, 'size': size, 'is_large': size > 5 * 1024 * 1024})  # >5MB视为大文件
    
    large_files = [f for f in file_info if f['is_large']]
    small_files = [f for f in file_info if not f['is_large']]
    
    # 大文件顺序读（fastexcel内部已多线程），小文件才并行
    max_workers = min(4, len(small_files)) if small_files else 1
    use_parallel = len(small_files) > 1 and max_workers > 1
    
    if large_files:
        log_func(f"  📁 检测到 {len(large_files)} 个大文件 (>5MB)，将顺序读取")
    if use_parallel:
        log_func(f"🚀 开启并行读取模式 ({max_workers} 个线程) 处理 {len(small_files)} 个小文件...")
    
    # 耗时统计
    time_read = 0
    time_convert = 0
    time_save = 0
    total_rows = 0
    
    # 使用Polars DataFrame列表存储（Arrow零拷贝）
    all_dfs = []
    all_columns = set()
    processed_count = 0
    
    # ========== 读取文件 ==========
    t_read_start = time.time()
    
    # 1. 先顺序读取所有大文件（fastexcel内部已多线程）
    for f_info in large_files:
        file_path = f_info['path']
        fname = os.path.basename(file_path)
        
        if stop_event and stop_event.is_set():
            log_func(">>> 用户强制停止任务！")
            return False, "任务已终止。"
        
        dfs, columns, count, read_time = read_excel_with_polars(file_path, fname, log_func)
        
        if dfs:
            all_dfs.extend(dfs)
            all_columns.update(columns)
            processed_count += count
            total_rows += sum(df.height for df in dfs)
            time_read += read_time
        else:
            log_func(f"  ⚠ {fname}: 大文件读取失败，尝试openpyxl...")
            try:
                temp_wb = load_workbook(file_path, read_only=False, data_only=True)
                ws = temp_wb.active
                file_data = []
                headers = None
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:
                        headers = [str(cell).strip() if cell is not None else f"column_{j}" for j, cell in enumerate(row)]
                        continue
                    row_data = list(row)
                    if any(cell is not None for cell in row_data):
                        row_dict = {}
                        for j, cell in enumerate(row_data):
                            col_name = headers[j] if j < len(headers) else f"column_{j}"
                            row_dict[col_name] = cell
                        row_dict['_来源工作簿'] = fname
                        row_dict['_来源工作表'] = ws.title
                        file_data.append(row_dict)
                if file_data:
                    fallback_df = pl.DataFrame(file_data, strict=False)
                    all_dfs.append(fallback_df)
                    all_columns.update(headers or [])
                    processed_count += 1
                    total_rows += len(file_data)
                    log_func(f"  ✓ {fname}: openpyxl备选成功 {len(file_data)} 行")
                temp_wb.close()
            except Exception as e2:
                log_func(f"  ✗ {fname}: openpyxl备选也失败")
    
    # 2. 并行读取所有小文件
    if use_parallel:
        log_func(f"  🚀 并行读取 {len(small_files)} 个小文件...")
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_file = {
                executor.submit(read_excel_parallel, f['path'], os.path.basename(f['path']), None): f['path']
                for f in small_files
            }
            
            for future in as_completed(future_to_file):
                if stop_event and stop_event.is_set():
                    log_func(">>> 用户强制停止任务！")
                    return False, "任务已终止。"
                
                file_path = future_to_file[future]
                fname = os.path.basename(file_path)
                
                try:
                    result = future.result()
                    
                    if result['dfs']:
                        all_dfs.extend(result['dfs'])
                        all_columns.update(result['columns'])
                        processed_count += result['count']
                        total_rows += result['rows']
                        time_read += result['elapsed']
                        log_func(f"  ✓ {result['fname']}: {result['rows']} 行 ({result['elapsed']:.2f}s)")
                    else:
                        log_func(f"  ⚠ {fname}: 小文件读取失败，尝试openpyxl...")
                        try:
                            temp_wb = load_workbook(file_path, read_only=False, data_only=True)
                            ws = temp_wb.active
                            file_data = []
                            headers = None
                            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                                if row_idx == 1:
                                    headers = [str(cell).strip() if cell is not None else f"column_{j}" for j, cell in enumerate(row)]
                                    continue
                                row_data = list(row)
                                if any(cell is not None for cell in row_data):
                                    row_dict = {}
                                    for j, cell in enumerate(row_data):
                                        col_name = headers[j] if j < len(headers) else f"column_{j}"
                                        row_dict[col_name] = cell
                                    row_dict['_来源工作簿'] = fname
                                    row_dict['_来源工作表'] = ws.title
                                    file_data.append(row_dict)
                            if file_data:
                                fallback_df = pl.DataFrame(file_data, strict=False)
                                all_dfs.append(fallback_df)
                                all_columns.update(headers or [])
                                processed_count += 1
                                total_rows += len(file_data)
                                log_func(f"  ✓ {fname}: openpyxl备选成功 {len(file_data)} 行")
                            temp_wb.close()
                        except Exception as e2:
                            log_func(f"  ✗ {fname}: openpyxl备选也失败")
                            
                except Exception as e:
                    log_func(f"  ✗ {fname}: 读取异常 - {e}")
    elif small_files:
        # 没有并行但有小文件，顺序读
        for f_info in small_files:
            file_path = f_info['path']
            fname = os.path.basename(file_path)
            
            if stop_event and stop_event.is_set():
                log_func(">>> 用户强制停止任务！")
                return False, "任务已终止。"
            
            dfs, columns, count, read_time = read_excel_with_polars(file_path, fname, log_func)
            
            if dfs:
                all_dfs.extend(dfs)
                all_columns.update(columns)
                processed_count += count
                total_rows += sum(df.height for df in dfs)
                time_read += read_time
    
    time_read = time.time() - t_read_start
    log_func(f"  📊 文件读取完成: 共 {total_rows} 行 (耗时 {time_read:.2f}s)")
    
    if not all_dfs:
        return False, "未提取到任何数据。"
    
    try:
        log_func("正在汇总数据...")
        log_func(f"  共读取 {total_rows} 行数据，正在合并...")
        
        # ========== Arrow零拷贝合并（驯服版） ==========
        t_convert_start = time.time()
        
        try:
            # 🔧 驯服Arrow的关键：先规范化所有DataFrame的schema
            # 收集所有出现过的列
            all_cols = set()
            for df in all_dfs:
                all_cols.update(df.columns)
            
            # 确保包含来源列
            all_cols.add('_来源工作簿')
            all_cols.add('_来源工作表')
            
            # 🔧 修复：保持第一个文件的列顺序（作为基准顺序）
            if all_dfs:
                # 以第一个DataFrame的列顺序为基准
                first_df_cols = list(all_dfs[0].columns)
                # 将其他列按首次出现的顺序追加
                seen = set(first_df_cols)
                for df in all_dfs[1:]:
                    for col in df.columns:
                        if col not in seen and col in all_cols:
                            first_df_cols.append(col)
                            seen.add(col)
                # 只保留目标列
                target_columns = [c for c in first_df_cols if c in all_cols]
                # 确保来源列在最后
                if '_来源工作簿' in all_cols and '_来源工作簿' not in target_columns:
                    target_columns.append('_来源工作簿')
                if '_来源工作表' in all_cols and '_来源工作表' not in target_columns:
                    target_columns.append('_来源工作表')
            else:
                target_columns = sorted(list(all_cols))
            
            # 规范化所有DataFrame，补全缺失的列
            normalized_dfs = normalize_dataframes_for_concat(all_dfs, target_columns, log_func)
            
            # 现在可以安全合并了！所有DataFrame都有相同的列结构
            combined_df = pl.concat(normalized_dfs, how="vertical_relaxed")
            time_convert = time.time() - t_convert_start
            log_func(f"  ✓ Arrow零拷贝合并完成: {combined_df.height} 行 ({time_convert:.2f}s)")
            
        except Exception as e:
            log_func(f"  ⚠ Arrow合并失败，回退到字典列表模式: {e}")
            # 回退方案：将DataFrame转换为字典列表
            all_data = []
            for df in all_dfs:
                rows = df.to_dicts()
                all_data.extend(rows)
            
            # 使用pandas处理
            pandas_df = pd.DataFrame(all_data)
            time_convert = time.time() - t_convert_start
            
            # 智能处理空值和类型
            for col in pandas_df.columns:
                pandas_df[col] = pandas_df[col].replace({None: pd.NA, 'None': pd.NA, 'nan': pd.NA, '': pd.NA})
                try:
                    non_null = pandas_df[col].dropna()
                    if len(non_null) > 0:
                        def is_number(x):
                            if pd.isna(x):
                                return False
                            try:
                                float(x)
                                return True
                            except:
                                return False
                        
                        if all(is_number(x) for x in non_null):
                            try:
                                pandas_df[col] = pd.to_numeric(pandas_df[col])
                            except (ValueError, TypeError):
                                pass
                except:
                    pass
            
            # 列处理
            source_cols = [c for c in pandas_df.columns if c not in ['_来源工作簿', '_来源工作表']]
            
            if extract_mode == 'specific':
                valid_cols = []
                for exact in exact_cols:
                    found = False
                    for col in source_cols:
                        if exact.lower() == col.lower():
                            if col not in valid_cols:
                                valid_cols.append(col)
                            found = True
                            break
                        if not found and is_fuzzy_match(exact, col, fuzzy_threshold):
                            if col not in valid_cols:
                                valid_cols.append(col)
                            found = True
                            break
                remaining = [c for c in source_cols if c not in valid_cols]
                final_cols = valid_cols + remaining
            else:
                final_cols = source_cols
            
            final_cols = final_cols + ['_来源工作簿', '_来源工作表']
            final_cols = [c for c in final_cols if c in pandas_df.columns]
            combined_df = pandas_df[final_cols]
            combined_polars = None  # 标记为pandas格式
        
        # ========== 列筛选 ==========
        log_func("  正在进行列筛选...")
        
        # 判断是Polars还是pandas格式
        if isinstance(combined_df, pl.DataFrame):
            # Polars格式处理
            source_cols = [c for c in combined_df.columns if c not in ['_来源工作簿', '_来源工作表']]
            
            if extract_mode == 'specific':
                valid_cols = []
                for exact in exact_cols:
                    found = False
                    for col in source_cols:
                        if exact.lower() == col.lower():
                            if col not in valid_cols:
                                valid_cols.append(col)
                            found = True
                            break
                        if not found and is_fuzzy_match(exact, col, fuzzy_threshold):
                            if col not in valid_cols:
                                valid_cols.append(col)
                            found = True
                            break
                remaining = [c for c in source_cols if c not in valid_cols]
                final_cols = valid_cols + remaining
            else:
                final_cols = source_cols
            
            final_cols = final_cols + ['_来源工作簿', '_来源工作表']
            final_cols = [c for c in final_cols if c in combined_df.columns]
            final_df = combined_df.select(final_cols)
            
            # ========== 保存数据 ==========
            t_save_start = time.time()
            
            if is_csv:
                # CSV极速保存 - 直接使用Polars写CSV（零拷贝）
                log_func(f"  🚀 Polars直接写CSV中...")
                final_df.write_csv(save_path, include_header=True, separator=',')
                time_save = time.time() - t_save_start
                log_func(f"  ✓ CSV保存完成! 耗时 {time_save:.2f}s")
                final_rows = final_df.height
            else:
                # Excel保存 - 需要转换为pandas
                log_func(f"  转换为pandas准备保存Excel...")
                pandas_df = final_df.to_pandas()
                
                if XLSXWRITER_AVAILABLE:
                    log_func(f"  🚀 使用xlsxwriter引擎保存...")
                    pandas_df.to_excel(save_path, index=False, na_rep='', engine='xlsxwriter')
                else:
                    pandas_df.to_excel(save_path, index=False, na_rep='')
                    log_func("  💡 提示: 安装xlsxwriter可提升保存速度 (pip install xlsxwriter)")
                
                time_save = time.time() - t_save_start
                final_rows = len(pandas_df)
        else:
            # Pandas格式处理（回退方案）
            source_cols = [c for c in combined_df.columns if c not in ['_来源工作簿', '_来源工作表']]
            
            if extract_mode == 'specific':
                valid_cols = []
                for exact in exact_cols:
                    found = False
                    for col in source_cols:
                        if exact.lower() == col.lower():
                            if col not in valid_cols:
                                valid_cols.append(col)
                            found = True
                            break
                        if not found and is_fuzzy_match(exact, col, fuzzy_threshold):
                            if col not in valid_cols:
                                valid_cols.append(col)
                            found = True
                            break
                remaining = [c for c in source_cols if c not in valid_cols]
                final_cols = valid_cols + remaining
            else:
                final_cols = source_cols
            
            final_cols = final_cols + ['_来源工作簿', '_来源工作表']
            final_cols = [c for c in final_cols if c in combined_df.columns]
            final_pandas_df = combined_df[final_cols]
            
            t_save_start = time.time()
            
            if is_csv:
                log_func(f"  🚀 CSV极速保存中...")
                final_pandas_df.to_csv(save_path, index=False, encoding='utf-8-sig')
                time_save = time.time() - t_save_start
                log_func(f"  ✓ CSV保存完成! 耗时 {time_save:.2f}s")
            else:
                if XLSXWRITER_AVAILABLE:
                    log_func(f"  🚀 使用xlsxwriter引擎保存...")
                    final_pandas_df.to_excel(save_path, index=False, na_rep='', engine='xlsxwriter')
                else:
                    final_pandas_df.to_excel(save_path, index=False, na_rep='')
                    log_func("  💡 提示: 安装xlsxwriter可提升保存速度 (pip install xlsxwriter)")
                time_save = time.time() - t_save_start
            
            final_rows = len(final_pandas_df)
        
        # ========== 性能统计 ==========
        total_time = time_read + time_convert + time_save
        log_func(f"  性能统计:")
        log_func(f"    读取阶段: {time_read:.2f}s ({time_read/total_time*100:.1f}%)")
        log_func(f"    转换阶段: {time_convert:.2f}s ({time_convert/total_time*100:.1f}%)")
        log_func(f"    保存阶段: {time_save:.2f}s ({time_save/total_time*100:.1f}%)")
        log_func(f"    总耗时: {total_time:.2f}s")
        
        engine_used = "fastexcel" if FASTEXCEL_AVAILABLE else "openpyxl"
        format_info = "CSV极速" if is_csv else "Excel"
        return True, f"Polars模式完成！使用{engine_used}引擎 + {format_info}，共提取 {final_rows} 行数据。\n保存至: {save_path}"
        
    except Exception as e:
        log_func(f"  处理异常: {e}")
        return False, f"保存失败: {e}"
        
    except Exception as e:
        return False, f"保存失败: {e}"

# ==================== 原有Pandas核心处理逻辑（已优化）====================
def process_with_pandas(files: List[str], target_sheets: List[str], 
                       extract_mode: str, exact_cols: List[str], 
                       fuzzy_cols: List[str], fuzzy_threshold: float,
                       save_path: str, log_func, stop_event=None,
                       is_csv: bool = False) -> Tuple[bool, str]:
    """使用Pandas引擎处理数据，支持CSV极速保存"""
    
    log_func("=== Pandas标准模式 ===")
    if is_csv:
        log_func("🚀 极速模式: CSV保存")
    
    # 【新增】耗时统计
    time_read = 0
    time_convert = 0
    time_save = 0
    total_rows = 0
    
    all_rows = []
    processed_count = 0
    
    for i, file_path in enumerate(files):
        # 检查中断
        if stop_event and stop_event.is_set():
            log_func(">>> 用户强制停止任务！")
            return False, "任务已终止。"
        
        fname = os.path.basename(file_path)
        t_start = time.time()
        log_func(f"[{i+1}/{len(files)}] 正在读取: {fname}")
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            for ws in wb.worksheets:
                # Sheet循环里检查中断
                if stop_event and stop_event.is_set():
                    wb.close()
                    log_func(">>> 用户强制停止任务！")
                    return False, "任务已终止。"
                
                if target_sheets:
                    is_match = False
                    for target in target_sheets:
                        if is_fuzzy_match(target, ws.title, fuzzy_threshold):
                            is_match = True
                            break
                    if not is_match:
                        continue
                
                col_map = scan_header_and_map_columns(ws, extract_mode, exact_cols, fuzzy_cols, fuzzy_threshold)
                if not col_map:
                    continue
                
                for row_data in ws.iter_rows(values_only=True):
                    extracted_row = {}
                    has_valid_data = False
                    for col_idx, target_name in col_map.items():
                        if col_idx - 1 < len(row_data):
                            val = row_data[col_idx - 1]
                            extracted_row[target_name] = val
                            if val is not None:
                                has_valid_data = True
                    
                    if has_valid_data:
                        match_header_count = sum(1 for k, v in extracted_row.items() if str(v) == k)
                        if match_header_count > 0 and match_header_count > len(extracted_row) / 2:
                            continue
                        extracted_row['_来源工作簿'] = fname
                        extracted_row['_来源工作表'] = ws.title
                        all_rows.append(extracted_row)
                        total_rows += 1
            
            wb.close()
            processed_count += 1
            log_func(f"  ✓ 读取成功: 累计 {total_rows} 行 (本文件耗时 {time.time()-t_start:.2f}s)")
            
        except InvalidFileException as e:
            log_func(f"  文件格式异常: {e}")
            # 尝试修复文件
            if repair_excel_file(file_path, log_func):
                try:
                    wb = load_workbook(file_path, read_only=True, data_only=True)
                    # 重复读取逻辑...
                    wb.close()
                    processed_count += 1
                except Exception as retry_e:
                    log_func(f"  ✗ 修复后仍失败: {retry_e}")
        except Exception as e:
            log_func(f"  ✗ 读取失败: {e}")
        
        time_read += time.time() - t_start
    
    if not all_rows:
        return False, "未提取到任何数据。"
    
    try:
        t_save_start = time.time()
        log_func("正在汇总并保存...")
        
        log_func(f"  共读取 {total_rows} 行数据，正在转换...")
        
        t_convert_start = time.time()
        df = pd.DataFrame(all_rows)
        time_convert = time.time() - t_convert_start
        
        cols = [c for c in df.columns if c not in ['_来源工作簿', '_来源工作表']]
        
        if extract_mode == 'specific':
            sorted_cols = []
            for exact in exact_cols:
                if exact in cols:
                    sorted_cols.append(exact)
            remaining = [c for c in cols if c not in sorted_cols]
            final_cols = sorted_cols + remaining + ['_来源工作簿', '_来源工作表']
        else:
            final_cols = cols + ['_来源工作簿', '_来源工作表']
            
        df = df.reindex(columns=final_cols)
        
        t_save_start = time.time()
        
        # 保存数据
        if is_csv:
            # CSV极速保存模式
            log_func(f"  🚀 CSV极速保存中...")
            df.to_csv(save_path, index=False, encoding='utf-8-sig')
            time_save = time.time() - t_save_start
            log_func(f"  ✓ CSV保存完成! 耗时 {time_save:.2f}s")
        else:
            # Excel保存（使用xlsxwriter引擎）
            if XLSXWRITER_AVAILABLE:
                df.to_excel(save_path, index=False, engine='xlsxwriter')
            else:
                df.to_excel(save_path, index=False)
            time_save = time.time() - t_save_start
        
        # 【新增】打印详细耗时
        total_time = time_read + time_convert + time_save
        log_func(f"  耗时统计:")
        log_func(f"    读取阶段: {time_read:.2f}s ({time_read/total_time*100:.1f}%)")
        log_func(f"    转换阶段: {time_convert:.2f}s ({time_convert/total_time*100:.1f}%)")
        log_func(f"    保存阶段: {time_save:.2f}s ({time_save/total_time*100:.1f}%)")
        log_func(f"    总耗时: {total_time:.2f}s")
        
        format_info = "CSV极速" if is_csv else "Excel"
        return True, f"Pandas模式完成！{format_info}格式，共提取 {len(df)} 行数据。\n保存至: {save_path}"
        
    except Exception as e:
        return False, f"保存失败: {e}"

# ==================== 统一入口函数 ====================
def core_process(source_path: str, is_folder: bool, recursive: bool, 
                target_sheets: List[str], extract_mode: str, 
                exact_cols: List[str], fuzzy_cols: List[str], 
                fuzzy_threshold: float, save_path: str, 
                log_func, stop_event=None,
                use_polars: bool = False,
                is_csv: bool = False) -> Tuple[bool, str]:
    """核心处理函数 - 支持双引擎，支持CSV极速保存"""
    
    log_func("=== 任务开始 ===")
    
    # 显示保存格式信息
    if is_csv:
        log_func("🚀 极速模式: 保存为CSV (比Excel快5-10倍)")
    else:
        log_func(f"保存为: {os.path.basename(save_path)}")
    
    files = get_files_to_process(source_path, is_folder, recursive)
    if not files:
        return False, "未找到有效的 Excel 文件。"
    
    # 根据设置选择处理引擎
    if use_polars and POLARS_AVAILABLE:
        return process_with_polars(
            files, target_sheets, extract_mode, exact_cols, 
            fuzzy_cols, fuzzy_threshold, save_path, log_func, stop_event,
            is_csv=is_csv
        )
    else:
        return process_with_pandas(
            files, target_sheets, extract_mode, exact_cols, 
            fuzzy_cols, fuzzy_threshold, save_path, log_func, stop_event,
            is_csv=is_csv
        )

# ==================== 界面层 ====================

class ColumnExtractorModule:
    def __init__(self):
        self.name = "Excel数据提取"
        self.input_path = ""
        self.app = None
        self.module_index = None
        
    def render(self, parent_frame):
        for widget in parent_frame.winfo_children():
            widget.destroy()

        STYLE_LABEL = {"font": ("Microsoft YaHei", 14), "text_color": "#333"}
        STYLE_ENTRY = {"height": 36, "border_color": "#D0D0D0", "border_width": 1, "fg_color": "#FAFAFA", "text_color": "#333"}
        STYLE_BTN_BLUE = {"height": 36, "fg_color": "#F0F5FA", "text_color": "#0984e3", "hover_color": "#E1EBF5"}
        
        scroll = ctk.CTkScrollableFrame(parent_frame, fg_color="transparent", scrollbar_button_color="#E0E0E0", scrollbar_button_hover_color="#D0D0D0")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(scroll, text="Excel数据提取器", font=("Microsoft YaHei", 24, "bold"), text_color="#333").pack(anchor="w", padx=20, pady=(10, 20))

        # --- 1. 数据源 ---
        frame_src = ctk.CTkFrame(scroll, fg_color="white", corner_radius=8, border_width=1, border_color="#E5E5E5")
        frame_src.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(frame_src, text="1. 数据源选择", font=("Microsoft YaHei", 15, "bold"), text_color="#0984e3").pack(anchor="w", padx=15, pady=10)
        
        self.var_source_type = ctk.StringVar(value="folder")
        f_type = ctk.CTkFrame(frame_src, fg_color="transparent")
        f_type.pack(fill="x", padx=15, pady=5)
        ctk.CTkRadioButton(f_type, text="处理文件夹 (批量)", variable=self.var_source_type, value="folder", text_color="#333", command=self.toggle_source_ui).pack(side="left", padx=(0, 20))
        ctk.CTkRadioButton(f_type, text="处理单个文件", variable=self.var_source_type, value="file", text_color="#333", command=self.toggle_source_ui).pack(side="left")

        f_path = ctk.CTkFrame(frame_src, fg_color="transparent")
        f_path.pack(fill="x", padx=15, pady=10)
        self.entry_src = ctk.CTkEntry(f_path, placeholder_text="请选择路径...", **STYLE_ENTRY)
        self.entry_src.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ctk.CTkButton(f_path, text="浏览...", width=100, command=self.select_source, **STYLE_BTN_BLUE).pack(side="left")

        self.check_recursive = ctk.CTkCheckBox(frame_src, text="递归遍历子文件夹 (仅文件夹模式有效)", text_color="#555", font=("Microsoft YaHei", 12))
        self.check_recursive.pack(anchor="w", padx=15, pady=(0, 15))

        # --- 2. 规则设置 ---
        frame_rule = ctk.CTkFrame(scroll, fg_color="white", corner_radius=8, border_width=1, border_color="#E5E5E5")
        frame_rule.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(frame_rule, text="2. 提取规则设置", font=("Microsoft YaHei", 15, "bold"), text_color="#0984e3").pack(anchor="w", padx=15, pady=10)

        f_sheet = ctk.CTkFrame(frame_rule, fg_color="transparent")
        f_sheet.pack(fill="x", padx=15, pady=5)
        ctk.CTkLabel(f_sheet, text="指定工作表：", width=100, anchor="e", **STYLE_LABEL).pack(side="left")
        self.entry_sheets = ctk.CTkEntry(f_sheet, placeholder_text="可选，留空则提取所有Sheet。逗号分隔", **STYLE_ENTRY)
        self.entry_sheets.pack(side="left", fill="x", expand=True)

        f_col_mode = ctk.CTkFrame(frame_rule, fg_color="transparent")
        f_col_mode.pack(fill="x", padx=15, pady=15)
        ctk.CTkLabel(f_col_mode, text="提取列模式：", width=100, anchor="e", **STYLE_LABEL).pack(side="left")
        self.switch_col_mode = ctk.CTkSwitch(f_col_mode, text="启用「指定列提取」", command=self.toggle_col_ui, text_color="#333", font=("Microsoft YaHei", 13))
        self.switch_col_mode.pack(side="left", padx=10)
        ctk.CTkLabel(f_col_mode, text="(关闭则自动提取所有发现的列)", text_color="gray", font=("Microsoft YaHei", 12)).pack(side="left")

        self.frame_col_detail = ctk.CTkFrame(frame_rule, fg_color="#F8F9FA", corner_radius=6)
        
        f_exact = ctk.CTkFrame(self.frame_col_detail, fg_color="transparent")
        f_exact.pack(fill="x", padx=10, pady=10)
        ctk.CTkLabel(f_exact, text="精确匹配列：", width=100, anchor="e", **STYLE_LABEL).pack(side="left")
        self.entry_exact = ctk.CTkEntry(f_exact, placeholder_text="列名完全一致才提取，逗号分隔", **STYLE_ENTRY)
        self.entry_exact.pack(side="left", fill="x", expand=True)

        f_fuzzy = ctk.CTkFrame(self.frame_col_detail, fg_color="transparent")
        f_fuzzy.pack(fill="x", padx=10, pady=(10, 5))
        ctk.CTkLabel(f_fuzzy, text="模糊匹配列：", width=100, anchor="e", **STYLE_LABEL).pack(side="left")
        self.entry_fuzzy = ctk.CTkEntry(f_fuzzy, placeholder_text="包含关键词即提取，逗号分隔", **STYLE_ENTRY)
        self.entry_fuzzy.pack(side="left", fill="x", expand=True)
        
        f_slider = ctk.CTkFrame(self.frame_col_detail, fg_color="transparent")
        f_slider.pack(fill="x", padx=10, pady=(0, 10))
        ctk.CTkLabel(f_slider, text="匹配严格度：", width=100, anchor="e", **STYLE_LABEL).pack(side="left")
        self.lbl_slider_val = ctk.CTkLabel(f_slider, text="0.3 (默认)", width=30, text_color="#0984e3", font=("Microsoft YaHei", 13, "bold"))
        self.lbl_slider_val.pack(side="right", padx=10)
        self.slider_fuzzy = ctk.CTkSlider(f_slider, from_=0.1, to=1.0, number_of_steps=90, command=self.update_slider_label)
        self.slider_fuzzy.set(0.3)
        self.slider_fuzzy.pack(side="left", fill="x", expand=True, padx=10)

        # --- 3. 引擎选择（fastexcel升级版）---
        frame_engine = ctk.CTkFrame(scroll, fg_color="white", corner_radius=8, border_width=1, border_color="#E5E5E5")
        frame_engine.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(frame_engine, text="3. 处理引擎选择", font=("Microsoft YaHei", 15, "bold"), text_color="#0984e3").pack(anchor="w", padx=15, pady=10)
        
        f_engine = ctk.CTkFrame(frame_engine, fg_color="transparent")
        f_engine.pack(fill="x", padx=15, pady=10)
        
        self.var_engine = ctk.StringVar(value="pandas")
        
        # Pandas选项
        self.radio_pandas = ctk.CTkRadioButton(
            f_engine, 
            text="Pandas (稳定兼容)", 
            variable=self.var_engine, 
            value="pandas",
            text_color="#333",
            command=self.toggle_engine_info
        )
        self.radio_pandas.pack(side="left", padx=(0, 20))
        
        # Polars选项（如果可用）
        if POLARS_AVAILABLE:
            self.radio_polars = ctk.CTkRadioButton(
                f_engine, 
                text="Polars + fastexcel (🚀极速)", 
                variable=self.var_engine, 
                value="polars",
                text_color="#E74C3C",  # 红色突出显示高性能
                command=self.toggle_engine_info
            )
            self.radio_polars.pack(side="left", padx=(0, 20))
            
            # fastexcel状态提示
            if FASTEXCEL_AVAILABLE:
                ctk.CTkLabel(
                    f_engine, 
                    text="⚡ fastexcel已安装 - Rust级别速度", 
                    text_color="#27AE60", 
                    font=("Microsoft YaHei", 11)
                ).pack(side="left")
            else:
                ctk.CTkLabel(
                    f_engine, 
                    text="💡 推荐安装fastexcel: pip install fastexcel", 
                    text_color="#F39C12", 
                    font=("Microsoft YaHei", 11)
                ).pack(side="left")
        else:
            ctk.CTkLabel(
                f_engine, 
                text="💡 Polars未安装 (pip install polars pyarrow fastexcel)", 
                text_color="#999999", 
                font=("Microsoft YaHei", 11)
            ).pack(side="left")
        
        # 引擎说明
        self.lbl_engine_info = ctk.CTkLabel(
            frame_engine,
            text="Pandas: 稳定兼容，适合所有场景。处理万行数据约需10-30秒。",
            text_color="#666666",
            font=("Microsoft YaHei", 11)
        )
        self.lbl_engine_info.pack(anchor="w", padx=15, pady=(0, 15))

        self.btn_run = ctk.CTkButton(scroll, text="开始执行提取", height=50, fg_color="#007AFF", font=("Microsoft YaHei", 16, "bold"), command=self.run_task)
        self.btn_run.pack(fill="x", padx=20, pady=20)

        ctk.CTkLabel(scroll, text="执行日志", font=("Microsoft YaHei", 14, "bold"), text_color="#333").pack(anchor="w", padx=20, pady=(0, 5))
        self.textbox = ctk.CTkTextbox(scroll, height=180, fg_color="#FAFAFA", border_width=1, border_color="#D0D0D0", text_color="#333")
        self.textbox.pack(fill="x", padx=20, pady=(0, 20))

        self.toggle_source_ui()
        self.toggle_col_ui()
        self.toggle_engine_info()

    def update_slider_label(self, value):
        val = round(value, 2)
        desc = ""
        if val >= 0.99:
            desc = "(精确)"
        elif val >= 0.8:
            desc = "(严格)"
        elif val <= 0.4:
            desc = "(宽松)"
        else:
            desc = "(标准)"
        self.lbl_slider_val.configure(text=f"{val} {desc}")

    def toggle_source_ui(self):
        mode = self.var_source_type.get()
        if mode == "file":
            self.check_recursive.configure(state="disabled")
            self.entry_src.configure(placeholder_text="请选择单个 .xlsx 文件")
        else:
            self.check_recursive.configure(state="normal")
            self.entry_src.configure(placeholder_text="请选择包含 Excel 的文件夹")

    def toggle_col_ui(self):
        if self.switch_col_mode.get() == 1:
            self.frame_col_detail.pack(fill="x", padx=15, pady=(0, 15))
        else:
            self.frame_col_detail.pack_forget()
    
    def toggle_engine_info(self):
        """更新引擎说明"""
        engine = self.var_engine.get()
        if engine == "pandas":
            self.lbl_engine_info.configure(
                text="Pandas: 稳定兼容，适合所有场景。处理万行数据约需10-30秒。"
            )
        else:
            if FASTEXCEL_AVAILABLE:
                self.lbl_engine_info.configure(
                    text="Polars + fastexcel: 🚀 Rust级别极速，利用多核CPU并行处理。处理万行数据约需1-3秒。"
                )
            else:
                self.lbl_engine_info.configure(
                    text="Polars: 高性能模式，需要安装fastexcel (pip install fastexcel)。"
                )

    def select_source(self):
        mode = self.var_source_type.get()
        path = ""
        if mode == "file":
            path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        else:
            path = filedialog.askdirectory()
        if path:
            self.entry_src.delete(0, "end")
            self.entry_src.insert(0, path)
            self.input_path = path

    def log(self, msg):
        self.textbox.insert("end", msg + "\n")
        self.textbox.see("end")

    def run_task(self):
        src = self.entry_src.get().strip()
        is_folder = (self.var_source_type.get() == "folder")
        recursive = (self.check_recursive.get() == 1)
        sheet_str = self.entry_sheets.get().strip()
        target_sheets = [s.strip() for s in sheet_str.replace("，", ",").split(",") if s.strip()]
        extract_mode = "specific" if self.switch_col_mode.get() == 1 else "all"
        exact_cols = []
        fuzzy_cols = []
        fuzzy_thresh = self.slider_fuzzy.get()
        use_polars = (self.var_engine.get() == "polars" and POLARS_AVAILABLE)
        
        if extract_mode == "specific":
            e_str = self.entry_exact.get().strip()
            f_str = self.entry_fuzzy.get().strip()
            if e_str:
                exact_cols = [x.strip() for x in e_str.replace("，", ",").split(",") if x.strip()]
            if f_str:
                fuzzy_cols = [x.strip() for x in f_str.replace("，", ",").split(",") if x.strip()]
            if not exact_cols and not fuzzy_cols:
                messagebox.showwarning("提示", "请至少填写一项匹配规则。")
                return

        if not src or not os.path.exists(src):
            messagebox.showerror("错误", "路径不存在")
            return

        # 支持CSV和Excel两种保存格式
        save_file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="提取结果.xlsx",
            filetypes=[
                ("Excel文件 (*.xlsx)", "*.xlsx"),
                ("CSV文件 (*.csv)", "*.csv")
            ]
        )
        if not save_file:
            return

        # 判断保存格式
        is_csv = save_file.lower().endswith('.csv')
        save_format = "CSV" if is_csv else "Excel"

        self.btn_run.configure(state="disabled", text="正在处理...")
        self.textbox.delete("1.0", "end")
        
        # 申请红旗
        stop_event = None
        if hasattr(self, 'app') and self.app is not None:
            stop_event = self.app.register_task(self.module_index)
        
        def t():
            success, msg = core_process(
                src, is_folder, recursive, target_sheets,
                extract_mode, exact_cols, fuzzy_cols, fuzzy_thresh,
                save_file, self.log,
                stop_event=stop_event,
                use_polars=use_polars,
                is_csv=is_csv
            )
            self.log("-" * 30)
            self.log(msg)
            
            # 任务结束，通知主程序
            if hasattr(self, 'app') and self.app is not None:
                self.app.finish_task(self.module_index)
                
            self.btn_run.configure(state="normal", text="开始执行提取")
            if success:
                engine_mode = "Polars + fastexcel极速" if use_polars else "Pandas标准"
                messagebox.showinfo("成功", f"数据提取完成！\n模式: {engine_mode}\n格式: {save_format}")

        threading.Thread(target=t, daemon=True).start()
