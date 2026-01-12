import os
import sys
import threading
import csv
import difflib
import re
import urllib.parse
import openpyxl 
from openpyxl import Workbook, load_workbook
import customtkinter as ctk
from tkinter import filedialog, messagebox
from concurrent.futures import ThreadPoolExecutor, as_completed

# --- 1. 引擎加载区 ---
try:
    from python_calamine import CalamineWorkbook
except ImportError:
    CalamineWorkbook = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

# COM 接口支持检查
try:
    import win32com.client as win32
    HAS_COM = True
except ImportError:
    HAS_COM = False

# ==============================================================================
#                               核心逻辑区
# ==============================================================================

# --- A. 内容检索逻辑 ---

def clean_val(val):
    if val is None: return ""
    val_str = str(val)
    if val_str.startswith("<") and "Error" in val_str:
        return val_str
    s = val_str.strip()
    if s.endswith(".0"):
        try:
            float(s); s = s[:-2]
        except: pass
    return s

def is_match(content, keyword, threshold):
    c_str = content.lower()
    k_str = keyword.lower()
    if not c_str: return False
    if threshold >= 0.99: return k_str in c_str 
    return difflib.SequenceMatcher(None, k_str, c_str).ratio() >= threshold

# === 【修改点 1】增加 stop_event 参数 ===
def scan_values_rust(file_path, keywords, threshold, stop_event=None):
    """Rust 极速查值"""
    hits = []
    try:
        wb = CalamineWorkbook.from_path(file_path)
        for sheet_name in wb.sheet_names:
            # === 中断检测 ===
            if stop_event and stop_event.is_set(): return hits
            
            try:
                rows = wb.get_sheet_by_name(sheet_name).to_python(skip_empty_area=False)
            except: continue
            for r_idx, row in enumerate(rows):
                for c_idx, cell_value in enumerate(row):
                    val_str = clean_val(cell_value)
                    if not val_str: continue
                    for kw in keywords:
                        if is_match(val_str, kw, threshold):
                            col_letter = openpyxl.utils.get_column_letter(c_idx + 1)
                            hits.append({
                                "file": os.path.basename(file_path),
                                "pos": f"{sheet_name}!{col_letter}{r_idx+1}",
                                "val": val_str
                            })
                            break
    except: pass
    return hits

# === 【修改点 2】增加 stop_event 参数 ===
def scan_values_openpyxl(file_path, keywords, threshold, stop_event=None):
    """OpenPyXL 查值"""
    hits = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            # === 中断检测 ===
            if stop_event and stop_event.is_set(): 
                wb.close()
                return hits
                
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if not cell.value: continue
                    val_str = clean_val(cell.value)
                    if not val_str: continue
                    for kw in keywords:
                        if is_match(val_str, kw, threshold):
                            hits.append({
                                "file": os.path.basename(file_path),
                                "pos": f"{sheet_name}!{cell.coordinate}",
                                "val": val_str
                            })
                            break
        wb.close()
    except: pass
    return hits

# --- B. 外部链接管理逻辑 (保持不变，此处不需要中断) ---

import zipfile
import xml.etree.ElementTree as ET

def parse_worksheet_xml_for_formulas(file_path, sheet_name):
    """直接解析工作表的XML文件，提取所有公式"""
    formulas = []
    
    try:
        # Excel文件是ZIP格式
        with zipfile.ZipFile(file_path, 'r') as zf:
            # 尝试读取工作表的XML文件
            sheet_xml_path = f"xl/worksheets/{sheet_name.lower()}.xml"
            # 尝试不同的命名格式
            possible_paths = [
                f"xl/worksheets/{sheet_name}.xml",
                f"xl/worksheets/sheet{sheet_name}.xml",
                f"xl/worksheets/sheet{sheet_name.lower()}.xml",
            ]
            
            xml_content = None
            for path in possible_paths:
                try:
                    xml_content = zf.read(path)
                    print(f"[XML解析] 找到工作表XML: {path}")
                    break
                except KeyError:
                    continue
            
            if not xml_content:
                # 列出所有文件
                all_files = zf.namelist()
                print(f"[XML解析] ZIP文件中的所有文件:")
                for f in all_files:
                    if 'worksheets' in f or 'sheet' in f.lower():
                        print(f"  {f}")
                return formulas
            
            # 解析XML
            root = ET.fromstring(xml_content)
            
            # 定义命名空间
            ns = {
                'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
            }
            
            # 查找所有公式
            # 1. 普通公式 <f> 在 <c> 单元格中
            for cell in root.findall('.//main:c', ns):
                cell_ref = cell.get('r', '')
                formula_elem = cell.find('main:f', ns)
                if formula_elem is not None:
                    formula_text = formula_elem.text or ''
                    # 查找单元格的值
                    value_elem = cell.find('main:v', ns)
                    value_text = value_elem.text if value_elem is not None else None
                    
                    formulas.append({
                        'cell': cell_ref,
                        'formula': formula_text,
                        'value': value_text
                    })
                    print(f"[XML解析] 找到公式: {cell_ref} = {formula_text[:80]}")
            
            # 2. 数组公式 <f> 在 <matrixFormula> 中
            for matrix_formula in root.findall('.//main:formula/main:array', ns):
                # 这种格式可能需要特殊处理
                pass
            
            # 3. 共享公式
            for shared_formula in root.findall('.//main:formula/main:shared', ns):
                pass
                
    except Exception as e:
        print(f"[XML解析] 错误: {e}")
        import traceback
        traceback.print_exc()
    
    return formulas


def process_extracted_formula(formula, base_links, filename_to_idx, found_refs):
    """从公式中提取外部引用并记录到 found_refs 中"""
    print(f"[扫描调试] 处理公式: {formula}")
    
    refs_in_formula = []
    
    # 模式1: 匹配单引号包围的工作表名: '[N]SheetName'! 或 '[完整路径]SheetName'!
    # 例如: ='[1]换这里'!$A$2 或 ='[D:\path\[BB.xlsx]SheetName'!A1
    pattern1 = r"'\[([^\]]+)\]([^']+)'!"
    matches1 = re.finditer(pattern1, formula)
    for match in matches1:
        idx_str = match.group(1)
        sheet_part = match.group(2)
        
        # 判断是数字索引还是文件名
        if idx_str.isdigit():
            idx = int(idx_str)
        else:
            # 是文件名，查找对应的索引
            idx_lower = idx_str.lower()
            if idx_lower in filename_to_idx:
                idx = filename_to_idx[idx_lower]
            else:
                # 没找到对应的索引，跳过
                print(f"[扫描调试]   警告: 文件名 '{idx_str}' 未找到对应的索引，跳过")
                continue
        
        refs_in_formula.append((idx, sheet_part))
        print(f"[扫描调试]   模式1匹配: 索引={idx}, Sheet='{sheet_part}'")
    
    # 模式2: 匹配没有单引号的工作表名: [N]SheetName!
    # 例如: =[1]Sheet1!A1 或 =SUMIFS([3]Sheet1!$B:$B,...)
    # 注意：工作表名不能包含单引号、!、] 这些特殊字符
    pattern2 = r'\[(\d+)\]([^!\'\[\]]+)!'
    matches2 = re.finditer(pattern2, formula)
    for match in matches2:
        idx = int(match.group(1))
        sheet_part = match.group(2)
        # 避免重复：如果已经通过模式1匹配过相同的引用，就不再添加
        is_duplicate = False
        for existing_idx, existing_sheet in refs_in_formula:
            if existing_idx == idx and existing_sheet == sheet_part:
                is_duplicate = True
                break
        if not is_duplicate:
            refs_in_formula.append((idx, sheet_part))
            print(f"[扫描调试]   模式2匹配: 索引={idx}, Sheet='{sheet_part}'")
    
    # 记录所有找到的引用
    for idx, sheet_part in refs_in_formula:
        if idx not in found_refs:
            found_refs[idx] = {
                "target": base_links.get(idx, "Unknown"),
                "sheets": set()
            }
        found_refs[idx]["sheets"].add(sheet_part)
        print(f"[扫描调试]   -> 记录 Sheet='{sheet_part}'")

def extract_links_from_file(file_path):
    links_info = []
    try:
        # 使用 data_only=False 以便读取公式而不是计算结果
        wb = load_workbook(file_path, data_only=False, keep_vba=True, read_only=False)
        
        # 先收集所有外部链接定义（从 _external_links）
        base_links = {}
        if hasattr(wb, "_external_links"):
            for idx, link in enumerate(wb._external_links):
                target = "Unknown"
                try: target = link.file_link.target
                except: pass
                # 保存完整路径
                base_links[idx + 1] = target 
                print(f"[扫描调试] 外部链接定义 [{idx+1}]: {target}")
        
        # 建立 filename -> idx 的映射（用于从公式中推断索引）
        filename_to_idx = {}
        for idx, target in base_links.items():
            filename = os.path.basename(target)
            filename_to_idx[filename.lower()] = idx
            print(f"[扫描调试] 文件名映射: {filename} -> 索引 {idx}")
        
        found_refs = {}  # {idx: {target: xxx, sheets: [sheet1, sheet2, ...]}}
        
        print(f"[扫描调试] 开始扫描公式...")
        
        # 方法1: 直接解析XML文件（更可靠）
        print(f"[扫描调试] 开始XML直接解析...")
        try:
            with zipfile.ZipFile(file_path, 'r') as zf:
                # 首先读取workbook.xml获取正确的sheet顺序
                sheet_id_to_name = {}
                try:
                    workbook_xml = zf.read('xl/workbook.xml')
                    workbook_root = ET.fromstring(workbook_xml)
                    ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    
                    # 遍历sheets元素
                    for sheet_elem in workbook_root.findall('.//main:sheets/main:sheet', ns):
                        sheet_id = sheet_elem.get('sheetId')
                        name = sheet_elem.get('name')
                        if sheet_id and name:
                            sheet_id_to_name[int(sheet_id)] = name
                    print(f"[XML解析] 从workbook.xml读取到 {len(sheet_id_to_name)} 个工作表")
                    for sid, sname in sheet_id_to_name.items():
                        print(f"  SheetId={sid}: {sname}")
                except Exception as wb_err:
                    print(f"[XML解析] 读取workbook.xml失败: {wb_err}")
                
                # 获取所有工作表文件
                sheet_files = [f for f in zf.namelist() if f.startswith('xl/worksheets/') and f.endswith('.xml')]
                print(f"[扫描调试] 发现 {len(sheet_files)} 个工作表XML文件")
                
                for sheet_xml_path in sheet_files:
                    # 从文件名提取工作表编号（假设格式为 sheet1.xml, sheet2.xml 等）
                    sheet_xml_name = os.path.basename(sheet_xml_path)
                    sheet_num = re.search(r'sheet(\d+)\.xml', sheet_xml_name)
                    if sheet_num:
                        sheet_idx = int(sheet_num.group(1))
                        # 优先使用workbook.xml中的名称映射
                        if sheet_idx in sheet_id_to_name:
                            sheet_name = sheet_id_to_name[sheet_idx]
                        elif sheet_idx <= len(wb.sheetnames):
                            sheet_name = wb.sheetnames[sheet_idx - 1]
                        else:
                            sheet_name = f"Sheet{sheet_idx}"
                        print(f"[扫描调试] 处理工作表XML: {sheet_xml_path} -> {sheet_name}")
                    else:
                        sheet_name = sheet_xml_name
                    
                    # 读取并解析XML
                    xml_content = zf.read(sheet_xml_path)
                    root = ET.fromstring(xml_content)
                    ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    
                    # 查找所有单元格中的公式
                    for cell in root.findall('.//main:c', ns):
                        cell_ref = cell.get('r', '')
                        formula_elem = cell.find('main:f', ns)
                        if formula_elem is not None and formula_elem.text:
                            formula_text = formula_elem.text
                            print(f"[扫描调试] XML找到公式: {cell_ref} = {formula_text[:100]}")
                            process_extracted_formula(formula_text, base_links, filename_to_idx, found_refs)
        except Exception as xml_err:
            print(f"[扫描调试] XML解析失败: {xml_err}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"[扫描调试] 扫描工作表: {sheet_name}, 行数: {ws.max_row}, 列数: {ws.max_column}")
            
            # 额外检查：扫描所有公式相关的属性
            # 有些公式可能存储在数组公式中
            if hasattr(ws, 'array_formulas') and ws.array_formulas:
                print(f"[扫描调试] 发现 {len(ws.array_formulas)} 个数组公式")
                for cell_coord, formula_data in ws.array_formulas.items():
                    print(f"[扫描调试]   数组公式位置: {cell_coord}, 数据: {formula_data}")
                    # 如果formula_data包含公式字符串，尝试提取
                    if isinstance(formula_data, dict) and 'formula' in formula_data:
                        formula_str = formula_data['formula']
                        if isinstance(formula_str, str) and formula_str.startswith("="):
                            print(f"[扫描调试]   从数组公式提取: {formula_str}")
                            process_extracted_formula(formula_str, base_links, filename_to_idx, found_refs)
                    elif isinstance(formula_data, str) and formula_data.startswith("="):
                        print(f"[扫描调试]   从数组公式提取(字符串): {formula_data}")
                        process_extracted_formula(formula_data, base_links, filename_to_idx, found_refs)
            
            # 额外检查：扫描shared_formulas
            if hasattr(ws, 'shared_formulas') and ws.shared_formulas:
                print(f"[扫描调试] 发现 {len(ws.shared_formulas)} 个共享公式")
                for cell_coord, formula_data in ws.shared_formulas.items():
                    print(f"[扫描调试]   共享公式位置: {cell_coord}, 数据: {formula_data}")
                    if isinstance(formula_data, str) and formula_data.startswith("="):
                        print(f"[扫描调试]   从共享公式提取: {formula_data}")
                        process_extracted_formula(formula_data, base_links, filename_to_idx, found_refs)
            
            # 额外检查：扫描所有单元格的原始属性
            print(f"[扫描调试] 开始详细扫描所有单元格...")
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 1):
                for cell in row:
                    # 详细检查每个单元格的所有属性
                    cell_info = {
                        'row': row_idx,
                        'col': cell.column,
                        'coordinate': cell.coordinate,
                        'value': cell.value,
                        'type(cell.value)': type(cell.value).__name__,
                        'data_type': getattr(cell, 'data_type', None),
                    }
                    
                    # 检查所有可能包含公式的属性
                    for attr_name in ['formula', '_value', 'internal_value', 'value_text', '_formula']:
                        if hasattr(cell, attr_name):
                            try:
                                val = getattr(cell, attr_name)
                                if val is not None:
                                    cell_info[attr_name] = f"{type(val).__name__}: {str(val)[:100]}"
                            except:
                                pass
                    
                    # 打印单元格信息
                    print(f"[扫描调试] 单元格 {cell.coordinate}: value={cell.value}, type={type(cell.value).__name__}")
                    
                    # 检查是否有公式
                    has_formula = False
                    for attr_name in ['formula', '_formula']:
                        if hasattr(cell, attr_name):
                            formula_val = getattr(cell, attr_name)
                            if formula_val and isinstance(formula_val, str) and formula_val.startswith("="):
                                print(f"[扫描调试]   -> 发现公式({attr_name}): {formula_val}")
                                has_formula = True
                                break
                    
                    # 继续扫描单元格，但只对有公式的单元格做详细处理
                    if has_formula:
                        continue
                    
                    formula = None
                    
                    # 方式1: 直接检查 cell.value
                    if cell.value is not None:
                        val_str = str(cell.value)
                        if val_str.startswith("="):
                            formula = val_str
                            print(f"[扫描调试] 方式1获取公式: {formula}")
                    
                    # 方式2: 处理 ArrayFormula 等特殊类型
                    if not formula:
                        if hasattr(cell, 'data_type') and cell.data_type == 'f':
                            print(f"[扫描调试] 方式2检测到公式类型: cell.value={type(cell.value)}")
                            if isinstance(cell.value, str):
                                formula = cell.value
                                print(f"[扫描调试] 方式2获取公式: {formula}")
                            elif hasattr(cell.value, 'value'):
                                # ArrayFormula 对象的处理方式
                                val = cell.value.value
                                print(f"[扫描调试]   ArrayFormula.value={val}")
                                if isinstance(val, str) and val.startswith("="):
                                    formula = val
                                    print(f"[扫描调试] 方式2获取公式: {formula}")
                                elif isinstance(val, list) and len(val) > 0:
                                    # 数组公式的公式字符串可能在其他地方
                                    for item in val:
                                        if isinstance(item, str) and item.startswith("="):
                                            formula = item
                                            print(f"[扫描调试] 方式2获取公式(数组): {formula}")
                                            break
                    
                    # 方式3: 检查内部属性
                    if not formula:
                        for attr_name in ['_value', 'internal_value', 'value_text']:
                            if hasattr(cell, attr_name):
                                try:
                                    val = getattr(cell, attr_name)
                                    if val is not None and isinstance(val, str) and val.startswith("="):
                                        formula = val
                                        print(f"[扫描调试] 方式3({attr_name})获取公式: {formula}")
                                        break
                                except:
                                    pass
                    
                    # 方式4: 尝试从 cell.formula 属性获取
                    if not formula and hasattr(cell, 'formula'):
                        formula_val = cell.formula
                        if formula_val and isinstance(formula_val, str) and formula_val.startswith("="):
                            formula = formula_val
                            print(f"[扫描调试] 方式4(formula属性)获取公式: {formula}")
                    
                    if formula:
                        print(f"[扫描调试] 处理公式: {formula}")
                        process_extracted_formula(formula, base_links, filename_to_idx, found_refs)
        
        # 转换为列表格式
        processed_indices = set()
        for idx, info in found_refs.items():
            for sheet in info["sheets"]:
                links_info.append({
                    "index": idx, 
                    "target": info["target"], 
                    "sheet": sheet
                })
                processed_indices.add(idx)
                print(f"[扫描调试] 记录: 索引={idx}, Sheet='{sheet}', Target={info['target']}")
        
        # 添加未被引用的链接
        for idx, target in base_links.items():
            if idx not in processed_indices:
                links_info.append({"index": idx, "target": target, "sheet": "(未使用)"})
                print(f"[扫描调试] 未使用: 索引={idx}, Target={target}")
        
        print(f"[扫描调试] 共发现 {len(links_info)} 条记录")
        wb.close()
    except Exception as e:
        print(f"[扫描错误] {e}")
        import traceback
        traceback.print_exc()
        return None, str(e)
    return links_info, ""

class ExcelComEngine:
    def __init__(self, log_func):
        self.log = log_func
        self.app = None
        
    def start(self):
        if not HAS_COM:
            self.log("错误: 未检测到 pywin32，无法调用 Excel。请安装: pip install pywin32")
            return False
        try:
            try: self.app = win32.Dispatch("Excel.Application")
            except: self.app = win32.Dispatch("Ket.Application") 
            self.app.Visible = False
            self.app.DisplayAlerts = False 
            return True
        except Exception as e:
            self.log(f"启动 Excel 失败: {e}")
            return False

    def close(self):
        if self.app:
            try: self.app.Quit()
            except: pass
            self.app = None

    def process_file(self, file_path, updates, original_sheets=None):
        """处理文件，支持路径和Sheet名替换（支持同索引多Sheet替换）"""
        abs_path = os.path.abspath(file_path)
        try:
            wb = self.app.Workbooks.Open(abs_path, UpdateLinks=0)
            current_links = wb.LinkSources(1)
            if not current_links:
                wb.Close(SaveChanges=False)
                return False, "文件内未检测到有效链接源"
            
            # 调试：打印所有外部链接
            self.log(f"[调试] 检测到 {len(current_links)} 个外部链接:")
            for i, link in enumerate(current_links, 1):
                self.log(f"[调试]   [{i}] {link}")
            
            changed = 0
            sheet_changed = 0
            formula_found = 0
            
            # 数据结构调整：
            # updates 结构为 {f_path: {'rules': [{'old_path_key': '原文件', 'new_path': '新路径', 'old_sheet': '原Sheet', 'new_sheet': '新Sheet'}, ...]}}
            
            # 注意：我们不使用 wb.ChangeLink，因为它会验证新路径中是否包含原 Sheet 名
            # 而是直接遍历所有公式，找到匹配的外部引用，然后修改路径和 Sheet 名
            
            # 收集所有替换规则（按路径片段匹配）
            rules = updates.get('rules', [])  # 新结构：列表，包含每个替换规则
            
            if rules:
                self.log(f"[调试] 开始替换，共 {len(rules)} 条规则")
            else:
                self.log(f"[调试] 无替换规则")
            
            # 遍历所有单元格，修改公式
            for sheet in wb.Worksheets:
                for cell in sheet.UsedRange:
                    try:
                        formula = cell.Formula
                        if not formula or not formula.startswith("="):
                            continue
                        
                        formula_found += 1
                        new_formula = None
                        self.log(f"[调试] 检查公式[{formula_found}]: {formula}")
                        
                        # 尝试匹配完整路径格式（可在公式任意位置）：='path[filename]SheetName'!Range
                        # 支持：$A$1（单个单元格）、$A$1:$N$29（范围）、$B:$B（整列）、$1:$1（整行）
                        
                        # 定义替换回调函数
                        def replace_external_ref(match):
                            path_part = match.group(1)  # 路径
                            filename = match.group(2)   # 文件名
                            current_sheet_clean = match.group(3)  # Sheet名
                            range_ref = match.group(4)  # 范围引用（不带!）
                            
                            # 通过文件名查找对应的规则列表
                            filename_key = filename.lower()
                            
                            # 遍历所有规则，找到文件名和Sheet名都匹配的规则
                            matched_rule = None
                            for rule in rules:
                                if rule['old_path_key'] == filename_key and rule['old_sheet'] == current_sheet_clean:
                                    matched_rule = rule
                                    break
                            
                            self.log(f"[调试] 回调: filename='{filename_key}', current_sheet='{current_sheet_clean}', rules_count={len(rules)}, matched={matched_rule is not None}")
                            
                            # 如果没有找到匹配的规则，返回原样
                            if not matched_rule:
                                return match.group(0)
                            
                            # 获取新路径和新Sheet名
                            new_path = matched_rule['new_path']
                            new_sheet_name = matched_rule['new_sheet']
                            
                            # 处理路径
                            if new_path:
                                new_filename = os.path.basename(new_path)
                                new_path_part = os.path.dirname(new_path)
                                if new_path_part:
                                    new_path_part = new_path_part.rstrip('/\\') + '\\'
                                else:
                                    new_path_part = ''
                            else:
                                new_filename = filename
                                new_path_part = path_part
                            
                            # 处理 Sheet 名
                            if new_sheet_name:
                                final_sheet = new_sheet_name
                            else:
                                final_sheet = current_sheet_clean
                            
                            # 构造新的外部引用部分
                            if re.search(r'\s', final_sheet):
                                new_external_ref = f"'{new_path_part}[{new_filename}]'{final_sheet}'!{range_ref}"
                            else:
                                new_external_ref = f"'{new_path_part}[{new_filename}]{final_sheet}'!{range_ref}"
                            
                            return new_external_ref
                        
                        # 使用re.sub进行替换
                        new_formula = re.sub(
                            r"'([^'\[\]]+)\[([^\]']+)\]([^']*)'!(\$?[A-Z]+\$?\d*(?::\$?[A-Z]+\$?\d*)?)",
                            replace_external_ref,
                            formula
                        )
                        
                        # 如果有替换，更新公式
                        if new_formula != formula:
                            try:
                                cell.Formula2 = new_formula
                            except:
                                cell.Formula = new_formula
                            sheet_changed += 1
                            changed += 1
                            self.log(f"[调试]     已替换路径")
                        
                        # 如果完整路径不匹配，尝试简化格式
                        if new_formula is None:
                            # 尝试匹配简化格式（不带单引号）：=[N]SheetName!
                            match2 = re.match(rf"^=\[(\d+)\]([^'!\[\]]+)!", formula)
                            if match2:
                                idx_num = int(match2.group(1))
                                current_sheet_clean = match2.group(2)
                                
                                if 0 < idx_num <= len(current_links):
                                    link_path = current_links[idx_num - 1]
                                    # 从link_path中提取文件名作为key
                                    filename_key = os.path.basename(link_path).lower()
                                    self.log(f"[调试] 简化格式: idx={idx_num}, link_path='{link_path}', filename_key='{filename_key}'")
                                    
                                    # 遍历所有规则，找到文件名和Sheet名都匹配的规则
                                    matched_rule = None
                                    for rule in rules:
                                        if rule['old_path_key'] == filename_key and rule['old_sheet'] == current_sheet_clean:
                                            matched_rule = rule
                                            break
                                    
                                    if matched_rule:
                                        new_path = matched_rule['new_path']
                                        new_sheet_name = matched_rule['new_sheet']
                                        
                                        range_match = re.search(r'!(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)', formula)
                                        if range_match:
                                            range_ref = "!" + range_match.group(1)
                                        else:
                                            range_ref = "!A1"
                                        
                                        if new_path:
                                            new_filename = os.path.basename(new_path)
                                            new_path_part = os.path.dirname(new_path)
                                            if new_path_part:
                                                new_path_part = new_path_part.rstrip('/\\') + '\\'
                                            else:
                                                new_path_part = ''
                                            
                                            if new_sheet_name:
                                                final_sheet = new_sheet_name
                                            else:
                                                final_sheet = current_sheet_clean
                                            
                                            # 转换为完整格式
                                            if re.search(r'\s', final_sheet):
                                                new_formula = f"='{new_path_part}[{new_filename}]'{final_sheet}'{range_ref}"
                                            else:
                                                new_formula = f"='{new_path_part}[{new_filename}]{final_sheet}'{range_ref}"
                                        else:
                                            # 没有路径替换，保持简化格式
                                            if new_sheet_name:
                                                final_sheet = new_sheet_name
                                            else:
                                                final_sheet = current_sheet_clean
                                            
                                            if re.search(r'\s', final_sheet):
                                                new_formula = f"=[{idx_num}]'{final_sheet}'{range_ref}"
                                            else:
                                                new_formula = f"=[{idx_num}]{final_sheet}{range_ref}"
                                        
                                        if new_formula and new_formula != formula:
                                            try:
                                                cell.Formula2 = new_formula
                                            except:
                                                cell.Formula = new_formula
                                            sheet_changed += 1
                                            changed += 1
                                            self.log(f"[调试]     已替换(简化)[{idx_num}]: '{current_sheet_clean}' -> '{final_sheet}'")
                        
                        # 如果还是不匹配，尝试简化格式带单引号：=[N]'SheetName'!
                        if new_formula is None:
                            match3 = re.match(rf"^=\[(\d+)\]'([^']+)'!", formula)
                            if match3:
                                idx_num = int(match3.group(1))
                                current_sheet_clean = match3.group(2)
                                
                                if 0 < idx_num <= len(current_links):
                                    link_path = current_links[idx_num - 1]
                                    # 从link_path中提取文件名作为key
                                    filename_key = os.path.basename(link_path).lower()
                                    self.log(f"[调试] 简化格式(带引号): idx={idx_num}, link_path='{link_path}', filename_key='{filename_key}'")
                                    
                                    # 遍历所有规则，找到文件名和Sheet名都匹配的规则
                                    matched_rule = None
                                    for rule in rules:
                                        if rule['old_path_key'] == filename_key and rule['old_sheet'] == current_sheet_clean:
                                            matched_rule = rule
                                            break
                                    
                                    if matched_rule:
                                        new_path = matched_rule['new_path']
                                        new_sheet_name = matched_rule['new_sheet']
                                        
                                        range_match = re.search(r'!(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)', formula)
                                        if range_match:
                                            range_ref = "!" + range_match.group(1)
                                        else:
                                            range_ref = "!$A$1"
                                        
                                        if new_path:
                                            new_filename = os.path.basename(new_path)
                                            new_path_part = os.path.dirname(new_path)
                                            if new_path_part:
                                                new_path_part = new_path_part.rstrip('/\\') + '\\'
                                            else:
                                                new_path_part = ''
                                            
                                            if new_sheet_name:
                                                final_sheet = new_sheet_name
                                            else:
                                                final_sheet = current_sheet_clean
                                            
                                            new_formula = f"='{new_path_part}[{new_filename}]{final_sheet}'{range_ref}"
                                        else:
                                            if new_sheet_name:
                                                final_sheet = new_sheet_name
                                            else:
                                                final_sheet = current_sheet_clean
                                            
                                            new_formula = f"='[{idx_num}]{final_sheet}'{range_ref}"
                                        
                                        if new_formula and new_formula != formula:
                                            try:
                                                cell.Formula2 = new_formula
                                            except:
                                                cell.Formula = new_formula
                                            sheet_changed += 1
                                            changed += 1
                                            self.log(f"[调试]     已替换(简化引号)[{idx_num}]: '{current_sheet_clean}' -> '{final_sheet}'")
                        
                        # 最后尝试特殊格式：='[N]SheetName'!
                        if new_formula is None:
                            match4 = re.match(r"^='\[(\d+)\]'([^']+)'!", formula)
                            if match4:
                                idx_num = int(match4.group(1))
                                current_sheet_clean = match4.group(2)
                                
                                if 0 < idx_num <= len(current_links):
                                    link_path = current_links[idx_num - 1]
                                    # 从link_path中提取文件名作为key
                                    filename_key = os.path.basename(link_path).lower()
                                    self.log(f"[调试] 完整格式(=引号): idx={idx_num}, link_path='{link_path}', filename_key='{filename_key}'")
                                    
                                    # 遍历所有规则，找到文件名和Sheet名都匹配的规则
                                    matched_rule = None
                                    for rule in rules:
                                        if rule['old_path_key'] == filename_key and rule['old_sheet'] == current_sheet_clean:
                                            matched_rule = rule
                                            break
                                    
                                    if matched_rule:
                                        new_path = matched_rule['new_path']
                                        new_sheet_name = matched_rule['new_sheet']
                                        
                                        range_match = re.search(r'!(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)', formula)
                                        if range_match:
                                            range_ref = "!" + range_match.group(1)
                                        else:
                                            range_ref = "!$A$1"
                                        
                                        if new_path:
                                            new_filename = os.path.basename(new_path)
                                            new_path_part = os.path.dirname(new_path)
                                            if new_path_part:
                                                new_path_part = new_path_part.rstrip('/\\') + '\\'
                                            else:
                                                new_path_part = ''
                                            
                                            if new_sheet_name:
                                                final_sheet = new_sheet_name
                                            else:
                                                final_sheet = current_sheet_clean
                                            
                                            new_formula = f"='{new_path_part}[{new_filename}]{final_sheet}'{range_ref}"
                                        else:
                                            if new_sheet_name:
                                                final_sheet = new_sheet_name
                                            else:
                                                final_sheet = current_sheet_clean
                                            
                                            new_formula = f"='[{idx_num}]{final_sheet}'{range_ref}"
                                        
                                        if new_formula and new_formula != formula:
                                            try:
                                                cell.Formula2 = new_formula
                                            except:
                                                cell.Formula = new_formula
                                            sheet_changed += 1
                                            changed += 1
                                            self.log(f"[调试]     已替换(=单引号)[{idx_num}]: '{current_sheet_clean}' -> '{final_sheet}'")
                    
                    except Exception as cell_err:
                        continue
            
            self.log(f"[调试] 共检查 {formula_found} 个公式, 替换了 {sheet_changed} 个Sheet")
            
            total_changed = changed + sheet_changed
            if total_changed > 0:
                wb.Save()
                wb.Close()
                return True, f"成功更新 {changed} 个链接源, {sheet_changed} 个Sheet引用"
            else:
                wb.Close(SaveChanges=False)
                return True, "无变化 (可能新路径与原路径相同)"
        except Exception as e:
            return False, f"处理出错: {e}"

# ==============================================================================
#                               界面模块
# ==============================================================================

class keyWordSearchModule:
    def __init__(self):
        self.name = "内容检索&链接管理"
        self.selected_paths = [] 
        self.search_results = []
        self.entry_map_path = None
        self.col_widths_s = [220, 150, 450] 
        # self.app 会由 main.py 注入

    def render(self, parent_frame):
        for widget in parent_frame.winfo_children(): widget.destroy()

        main_scroll = ctk.CTkScrollableFrame(parent_frame, fg_color="transparent", scrollbar_button_color="#E0E0E0", scrollbar_button_hover_color="#D0D0D0")
        main_scroll.pack(fill="both", expand=True, padx=0, pady=0)

        title_frame = ctk.CTkFrame(main_scroll, fg_color="transparent")
        title_frame.pack(fill="x", padx=10, pady=(15, 5))
        ctk.CTkLabel(title_frame, text="🔍 内容检索 & 外部链接管理", font=("Microsoft YaHei", 20, "bold"), text_color="#333").pack(side="left")

        self.tabview = ctk.CTkTabview(main_scroll, fg_color="transparent", segmented_button_fg_color="#F0F0F0", segmented_button_selected_color="#0984e3", segmented_button_selected_hover_color="#0984e3", segmented_button_unselected_color="#E0E0E0", segmented_button_unselected_hover_color="#D6D6D6", text_color="#333", height=600)
        self.tabview.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        self.tab_search = self.tabview.add("内容/数值检索")
        self.tab_link = self.tabview.add("外部链接替换")

        self.render_search_tab(self.tab_search)
        self.render_link_tab(self.tab_link)

    # ------------------ Tab 1: 内容检索 ------------------
    def render_search_tab(self, parent):
        opt_frame = ctk.CTkFrame(parent, fg_color="white", corner_radius=6, border_width=1, border_color="#E0E0E0")
        opt_frame.pack(fill="x", padx=0, pady=5)

        row1 = ctk.CTkFrame(opt_frame, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=(15, 10))
        ctk.CTkButton(row1, text="+ 文件夹", command=self.add_folder, width=90, fg_color="#F0F5FF", text_color="#007AFF").pack(side="left", padx=5)
        ctk.CTkButton(row1, text="+ 文件", command=self.add_files, width=80, fg_color="#F0F5FF", text_color="#007AFF").pack(side="left", padx=5)
        self.lbl_count = ctk.CTkLabel(row1, text="未选择文件", text_color="#999")
        self.lbl_count.pack(side="left", padx=15)
        ctk.CTkButton(row1, text="清空", command=self.clear_selection, width=60, fg_color="transparent", text_color="#d63031").pack(side="right", padx=5)

        row2 = ctk.CTkFrame(opt_frame, fg_color="transparent")
        row2.pack(fill="x", padx=10, pady=(0, 10))
        self.entry_kw = ctk.CTkEntry(row2, placeholder_text="输入查找内容 (支持多内容，请用 | 分隔，如: 采购|销售|1,000)", height=40)
        self.entry_kw.pack(fill="x")

        row3 = ctk.CTkFrame(opt_frame, fg_color="transparent")
        row3.pack(fill="x", padx=15, pady=(0, 15))
        ctk.CTkLabel(row3, text="匹配度:", text_color="#333", width=60).pack(side="left")
        self.lbl_slider_val = ctk.CTkLabel(row3, text="1.0", text_color="#0984e3", width=30)
        self.lbl_slider_val.pack(side="left")
        self.slider = ctk.CTkSlider(row3, from_=0.1, to=1.0, number_of_steps=90, width=150, command=lambda v: self.lbl_slider_val.configure(text=f"{v:.1f}"))
        self.slider.set(1.0)
        self.slider.pack(side="left", padx=5)
        
        self.var_rust = ctk.BooleanVar(value=False)
        if CalamineWorkbook: ctk.CTkSwitch(row3, text="Rust极速模式", variable=self.var_rust, text_color="#555").pack(side="right")

        btn_row = ctk.CTkFrame(parent, fg_color="transparent")
        btn_row.pack(fill="x", padx=0, pady=10)
        self.btn_run_search = ctk.CTkButton(btn_row, text="开始检索", command=self.run_search, height=45, fg_color="#0984e3", font=("Microsoft YaHei", 15, "bold"))
        self.btn_run_search.pack(side="left", fill="x", expand=True, padx=(0, 5))
        ctk.CTkButton(btn_row, text="导出结果", command=self.export_search, height=45, fg_color="#00b894", font=("Microsoft YaHei", 15, "bold")).pack(side="right", fill="x", expand=True, padx=(5, 0))

        self.lbl_search_status = ctk.CTkLabel(parent, text="就绪", text_color="#666", anchor="w")
        self.lbl_search_status.pack(fill="x", padx=5, pady=(0, 5))

        res_container = ctk.CTkFrame(parent, fg_color="transparent")
        res_container.pack(fill="both", expand=True, padx=0, pady=5)
        header_grid = ctk.CTkFrame(res_container, fg_color="#E0E0E0", height=30, corner_radius=2)
        header_grid.pack(fill="x")
        headers = ["文件", "位置", "内容"]
        for i, h in enumerate(headers): ctk.CTkLabel(header_grid, text=h, width=self.col_widths_s[i], anchor="w", font=("Arial", 11, "bold")).pack(side="left", padx=10)

        self.res_frame = ctk.CTkFrame(res_container, fg_color="white", corner_radius=0)
        self.res_frame.pack(fill="both", expand=True)
        ctk.CTkLabel(self.res_frame, text="暂无结果", text_color="#CCC", height=50).pack()

    # ------------------ Tab 2: 外部链接管理 ------------------
    def render_link_tab(self, parent):
        step1 = ctk.CTkFrame(parent, fg_color="white", corner_radius=6, border_width=1, border_color="#E0E0E0")
        step1.pack(fill="x", padx=0, pady=10)
        s1_head = ctk.CTkFrame(step1, fg_color="transparent")
        s1_head.pack(fill="x", padx=15, pady=(15, 5))
        ctk.CTkLabel(s1_head, text="1. 扫描导出", font=("Microsoft YaHei", 14, "bold"), text_color="#333").pack(side="left")
        ctk.CTkLabel(s1_head, text="(需先在【内容检索】页选择文件)", text_color="#999", font=("Microsoft YaHei", 12)).pack(side="left", padx=10)
        ctk.CTkButton(step1, text="导出映射表 (Excel)", command=self.run_link_scan, height=40, fg_color="#0984e3", font=("Microsoft YaHei", 13, "bold")).pack(fill="x", padx=15, pady=(5, 15))

        step2 = ctk.CTkFrame(parent, fg_color="white", corner_radius=6, border_width=1, border_color="#E0E0E0")
        step2.pack(fill="x", padx=0, pady=0)
        s2_head = ctk.CTkFrame(step2, fg_color="transparent")
        s2_head.pack(fill="x", padx=15, pady=(15, 5))
        ctk.CTkLabel(s2_head, text="2. 替换执行", font=("Microsoft YaHei", 14, "bold"), text_color="#333").pack(side="left")
        ctk.CTkLabel(s2_head, text="(调用 Excel/WPS 进程)", text_color="#d63031", font=("Microsoft YaHei", 12)).pack(side="left", padx=10)
        
        f_in = ctk.CTkFrame(step2, fg_color="transparent")
        f_in.pack(fill="x", padx=15, pady=(0, 10))
        self.entry_map_path = ctk.CTkEntry(f_in, placeholder_text="选择映射表...", height=35)
        self.entry_map_path.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(f_in, text="浏览", width=70, command=self.select_map_file, height=35, fg_color="#F0F0F0", text_color="#333", hover_color="#DDD").pack(side="right", padx=(5,0))

        self.btn_run_com = ctk.CTkButton(step2, text="启动 Excel 执行替换", command=self.run_link_replace_com, height=40, fg_color="#d63031", hover_color="#b71c1c", font=("Microsoft YaHei", 13, "bold"))
        self.btn_run_com.pack(fill="x", padx=15, pady=(5, 15))

        ctk.CTkLabel(parent, text="执行日志:", text_color="#666", anchor="w", font=("Arial", 12, "bold")).pack(fill="x", padx=5, pady=(15, 5))
        self.log_box = ctk.CTkTextbox(parent, height=180, fg_color="white", border_width=1, border_color="#DDD", text_color="#333", font=("Consolas", 11))
        self.log_box.pack(fill="x", padx=0, pady=(0, 10))

    # ================= 交互逻辑 =================
    
    # ... (辅助函数保持不变) ...
    def add_folder(self):
        d = filedialog.askdirectory()
        if d: self.selected_paths.append(d); self.update_cnt()
    def add_files(self):
        fs = filedialog.askopenfilenames()
        if fs: self.selected_paths.extend(fs); self.update_cnt()
    def clear_selection(self):
        self.selected_paths = []; self.update_cnt()
    def update_cnt(self):
        self.lbl_count.configure(text=f"已选 {len(self.selected_paths)} 项")
    def select_map_file(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p: self.entry_map_path.delete(0, "end"); self.entry_map_path.insert(0, p)
    def log_link(self, msg):
        self.log_box.insert("end", msg + "\n"); self.log_box.see("end")
    def get_all_files(self):
        files = []
        for p in self.selected_paths:
            if os.path.isfile(p): files.append(p)
            elif os.path.isdir(p):
                for root, _, fs in os.walk(p):
                    for f in fs:
                        if not f.startswith("~$") and f.endswith((".xlsx", ".xlsm")): files.append(os.path.join(root, f))
        return list(set(files))

    # --- Tab 1 任务逻辑 ---
    def run_search(self):
        raw_kw = self.entry_kw.get()
        kws = [k.strip() for k in raw_kw.split('|') if k.strip()]
        if not self.selected_paths or not kws: return messagebox.showwarning("提示", "请选择文件并输入关键词")
        
        # 申请红旗
        stop_event = None
        if hasattr(self, 'app'): stop_event = self.app.register_task(self.module_index)
        self.btn_run_search.configure(state="disabled", text="搜索中...")
        
        for w in self.res_frame.winfo_children(): w.destroy()
        self.lbl_search_status.configure(text="正在初始化...")
        
        files = self.get_all_files()
        threshold = self.slider.get()
        use_rust = self.var_rust.get()
        
        def task():
            results = []
            total_files = len(files)
            aborted = False
            
            for i, f in enumerate(files):
                # === 中断检测 (文件级) ===
                if stop_event and stop_event.is_set():
                    self.lbl_search_status.configure(text=">>> 搜索已被强制终止。")
                    aborted = True
                    break
                    
                self.lbl_search_status.configure(text=f"正在搜索 ({i+1}/{total_files}): {os.path.basename(f)} ...")
                
                # 传入 stop_event 到核心函数
                if use_rust and CalamineWorkbook:
                    res = scan_values_rust(f, kws, threshold, stop_event=stop_event)
                else:
                    res = scan_values_openpyxl(f, kws, threshold, stop_event=stop_event)
                results.extend(res)
            
            if not aborted:
                self.search_results = results
                self.lbl_search_status.configure(text=f"搜索完成。共找到 {len(results)} 条结果。")
                # UI更新部分省略(保持不变)...
                # (为了简洁，这里省略 update_ui 内部代码，因为没有逻辑变更，只在 task 结尾调用)
                self.after_search_complete(results)
            
            # 销假 & 恢复按钮
            if hasattr(self, 'app'): self.app.finish_task(self.module_index)
            self.btn_run_search.configure(state="normal", text="开始检索")
            
        threading.Thread(target=task, daemon=True).start()

    def after_search_complete(self, results):
        # 辅助函数：更新搜索结果 UI
        for w in self.res_frame.winfo_children(): w.destroy()
        if not results: 
            ctk.CTkLabel(self.res_frame, text="未找到匹配项", text_color="#999").pack(pady=20)
            return
        limit = 10 # 限制显示条数避免卡顿
        for i, r in enumerate(results):
            if i >= limit:
                ctk.CTkLabel(self.res_frame, text=f"... 剩余 {len(results)-limit} 条请导出 ...", text_color="#d63031").pack(pady=5)
                break
            row = ctk.CTkFrame(self.res_frame, fg_color="transparent", height=30)
            row.pack(fill="x")
            row.pack_propagate(False)
            f1 = ctk.CTkFrame(row, fg_color="transparent", width=self.col_widths_s[0]); f1.pack(side="left", fill="y"); f1.pack_propagate(False)
            ctk.CTkLabel(f1, text=r['file'], anchor="w", font=("Arial", 11)).pack(side="left", padx=10)
            f2 = ctk.CTkFrame(row, fg_color="transparent", width=self.col_widths_s[1]); f2.pack(side="left", fill="y"); f2.pack_propagate(False)
            ctk.CTkLabel(f2, text=r['pos'], anchor="w", font=("Arial", 11), text_color="#00b894").pack(side="left")
            f3 = ctk.CTkFrame(row, fg_color="transparent", width=self.col_widths_s[2]); f3.pack(side="left", fill="y"); f3.pack_propagate(False)
            ctk.CTkLabel(f3, text=r['val'], anchor="w", font=("Arial", 11)).pack(side="left")
            ctk.CTkFrame(self.res_frame, fg_color="#F0F0F0", height=1).pack(fill="x")

    def export_search(self):
        if not self.search_results: return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="检索结果.xlsx", filetypes=[("Excel", "*.xlsx")])
        if path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.append(["文件", "位置", "内容"])
                for r in self.search_results:
                    val = str(r['val']) if r['val'] else ""
                    val = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)
                    ws.append([r['file'], r['pos'], val])
                wb.save(path)
                messagebox.showinfo("完成", "导出成功")
            except Exception as e:
                messagebox.showerror("失败", str(e))

    # --- Tab 2 任务逻辑 ---
    def run_link_scan(self):
        # 扫描很快，不加中断
        files = self.get_all_files()
        if not files: return messagebox.showwarning("提示", "请先在【内容检索】页签添加 Excel 文件")
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="外部链接映射表.xlsx")
        if not save_path: return

        self.log_box.delete("1.0", "end")
        self.log_link("开始扫描外部链接...")

        def task():
            try:
                wb_out = Workbook(); ws_out = wb_out.active
                ws_out.append(["文件全路径", "文件名", "链接索引([N])", "当前链接路径", "原引用Sheet名(参考)", "新链接路径(必填)", "新Sheet名(选填)"])
                count = 0
                for f in files:
                    links, err = extract_links_from_file(f)
                    if err: self.log_link(f"[错] {os.path.basename(f)}: {err}"); continue
                    if links:
                        for l in links: ws_out.append([f, os.path.basename(f), l['index'], l['target'], l['sheet'], "", ""])
                        count += 1
                        self.log_link(f"[扫描] {os.path.basename(f)} 发现 {len(links)} 个引用点")
                wb_out.save(save_path)
                self.log_link(f"-"*30)
                self.log_link(f"扫描结束。共发现含有链接的文件: {count} 个")
                self.entry_map_path.delete(0, "end"); self.entry_map_path.insert(0, save_path)
            except Exception as e:
                self.log_link(f"扫描导出失败: {e}")

        threading.Thread(target=task, daemon=True).start()

    def run_link_replace_com(self):
        map_file = self.entry_map_path.get()
        if not os.path.exists(map_file): return messagebox.showerror("错误", "找不到映射表文件")
        if not messagebox.askyesno("确认", "将启动 Excel 进程执行更改源。\n\n请确保：\n1. 所有 Excel 文件已关闭\n2. 已安装 pywin32\n\n是否继续？"): return

        # 申请红旗
        stop_event = None
        if hasattr(self, 'app'): stop_event = self.app.register_task(self.module_index)
        self.btn_run_com.configure(state="disabled", text="执行中...")

        self.log_link("\n启动 Excel 引擎...")
        
        def task():
            try:
                wb_map = load_workbook(map_file, data_only=True)
                ws_map = wb_map.active
                tasks = {} 
                for row in ws_map.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 6: continue
                    # 读取：当前链接路径(D列, 索引3)、新链接路径(E列, 索引5) 和 新Sheet名(F列, 索引6)
                    f_path, idx, current_path, original_sheet, new_path, new_sheet = row[0], row[2], row[3], row[4], row[5], row[6]
                    
                    # 判断是否有替换任务（新路径或新Sheet名非空）
                    has_path_update = new_path and str(new_path).strip()
                    has_sheet_update = new_sheet and str(new_sheet).strip()
                    
                    if not has_path_update and not has_sheet_update:
                        continue  # 跳过空行
                    
                    if f_path not in tasks:
                        tasks[f_path] = {}
                    
                    # 使用当前链接路径的一部分作为匹配key（取路径的末尾部分）
                    current_path_str = str(current_path).strip() if current_path else ""
                    if not current_path_str:
                        continue  # 跳过没有当前路径的行
                    
                    # 取路径的最后一部分作为key（提取文件名）
                    # 注意：用户可能填写完整路径或相对路径，需要统一处理
                    path_key = current_path_str.strip()
                    # 移除路径开头的反斜杠（Windows路径格式）
                    if path_key.startswith('\\'):
                        path_key = path_key[1:]
                    # 只取文件名部分
                    path_key = os.path.basename(path_key)
                    if not path_key:
                        path_key = current_path_str.strip()
                    path_key = path_key.lower()
                    
                    # 提取原Sheet名
                    old_sheet_clean = str(original_sheet).strip() if original_sheet and original_sheet != "(未使用)" else ""
                    
                    # 调试日志：显示本行的值
                    self.log_link(f"[调试] 处理行: path_key='{path_key}', has_path_update={has_path_update}, has_sheet_update={has_sheet_update}, old_sheet='{old_sheet_clean}'")
                    
                    # 只有当同时填写了"原引用Sheet名"时才处理
                    if old_sheet_clean:
                        if 'rules' not in tasks[f_path]:
                            tasks[f_path]['rules'] = []
                        
                        # 每行是一个完整的替换规则
                        rule = {
                            'old_path_key': path_key,  # 原文件（BB.xlsx）
                            'new_path': str(new_path).strip() if new_path else None,  # 新文件（可为None）
                            'old_sheet': old_sheet_clean,
                            'new_sheet': str(new_sheet).strip() if new_sheet else None
                        }
                        tasks[f_path]['rules'].append(rule)
                        self.log_link(f"[调试] 添加规则: 原文件='{path_key}', 新文件={rule['new_path']}, '{old_sheet_clean}' -> '{rule['new_sheet']}'")
                
                self.log_link(f"[调试] 解析任务: {tasks}")
            except Exception as e:
                self.log_link(f"映射表读取失败: {e}"); return

            if not tasks: self.log_link("未发现任务 (E列为空)"); return

            engine = ExcelComEngine(self.log_link)
            if not engine.start(): return
            
            success_cnt = 0
            for f_path, file_updates in tasks.items():
                # === 中断检测 ===
                if stop_event and stop_event.is_set():
                    self.log_link(">>> 用户强制终止！")
                    break

                if not os.path.exists(f_path): self.log_link(f"[跳过] 文件不存在: {f_path}"); continue
                
                self.log_link(f"正在处理: {os.path.basename(f_path)} ...")
                ok, msg = engine.process_file(f_path, file_updates)
                if ok: self.log_link(f"  -> {msg}"); success_cnt += 1
                else: self.log_link(f"  -> [失败] {msg}")
            
            engine.close()
            self.log_link(f"-"*30)
            self.log_link(f"任务结束。成功: {success_cnt}")
            
            # 销假 & 恢复按钮
            if hasattr(self, 'app'): self.app.finish_task(self.module_index)
            self.btn_run_com.configure(state="normal", text="启动 Excel 执行替换")

        threading.Thread(target=task, daemon=True).start()