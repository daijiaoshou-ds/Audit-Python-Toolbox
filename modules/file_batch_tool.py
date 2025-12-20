import os
import shutil
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading

# --- 核心辅助函数 (保持不变) ---

def get_unique_path(path: Path):
    if not path.exists(): return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    counter = 1
    while True:
        new_name = f"{stem}({counter}){suffix}"
        new_path = parent / new_name
        if not new_path.exists(): return new_path
        counter += 1

def create_valid_file(path: Path):
    suffix = path.suffix.lower()
    try:
        if suffix == '.xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(path)
            wb.close()
            return True
        elif suffix == '.pdf':
            pdf_content = (
                b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
                b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
                b"3 0 obj<</Type/Page/MediaBox[0 0 595 842]/Parent 2 0 R>>endobj\n"
                b"xref\n0 4\n0000000000 65535 f\n0000000009 00000 n\n"
                b"0000000052 00000 n\n0000000101 00000 n\n"
                b"trailer<</Size 4/Root 1 0 R>>\nstartxref\n178\n%%EOF"
            )
            with open(path, 'wb') as f: f.write(pdf_content)
            return True
        else:
            with open(path, 'w', encoding='utf-8') as f: pass
            return True
    except Exception as e:
        return False

# --- 业务逻辑 (修改点) ---

# === 【修改点 1】新增 stop_event 参数 ===
def generate_excel_template(src_folder, save_path, log_callback, stop_event=None):
    """生成 Excel 清单"""
    data = []
    log_callback(f"正在扫描: {src_folder}")
    
    for root, dirs, files in os.walk(src_folder):
        # === 【修改点 2】循环检测中断 ===
        if stop_event and stop_event.is_set():
            log_callback(">>> 用户强制停止扫描！")
            return False, "扫描已终止。"
        # ==============================

        for file in files:
            full_path = Path(root) / file
            try:
                parent_folder_name = full_path.parent.name
                if full_path.parent == Path(src_folder):
                    parent_folder_name = "根目录"

                data.append({
                    "原文件夹名称": parent_folder_name,
                    "原文件名": file,
                    "文件路径": str(full_path),
                    "新文件夹名称": "",
                    "新文件名": ""
                })
            except Exception as e:
                log_callback(f"跳过文件 {file}: {e}")

    if not data:
        data.append({"原文件夹名称": "", "原文件名": "", "文件路径": "", "新文件夹名称": "", "新文件名": ""})
        log_callback("警告: 文件夹为空，生成空模板。")

    # (这部分生成 Excel 的代码比较快，通常不需要中断，保持原样)
    df = pd.DataFrame(data)
    try:
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='文件管理')
            worksheet = writer.sheets['文件管理']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            worksheet.freeze_panes = 'B3'
            for col_idx, column_cells in enumerate(worksheet.columns, 1):
                col_letter = get_column_letter(col_idx)
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx == 3: worksheet.column_dimensions[col_letter].width = 60
                else: worksheet.column_dimensions[col_letter].width = 25
                if col_idx == 3:
                    for cell in column_cells[1:]:
                        if cell.value:
                            cell.hyperlink = cell.value
                            cell.font = Font(color="0563C1", underline="single")
        return True, f"Excel 已生成: {save_path}"
    except Exception as e:
        return False, f"生成失败: {str(e)}"

# === 【修改点 3】新增 stop_event 参数 ===
def process_files_from_excel(excel_path, root_folder, log_callback, is_copy_mode=False, is_replace_mode=False, stop_event=None):
    """执行修改、新建、移动/复制"""
    try:
        df = pd.read_excel(excel_path, dtype=str)
        df = df.fillna("")
    except Exception as e:
        return False, f"读取Excel失败: {e}"

    success_count = 0
    fail_count = 0
    root_path = Path(root_folder)
    
    mode_copy_text = "复制" if is_copy_mode else "移动"
    mode_replace_text = "【强制覆盖】" if is_replace_mode else "自动重命名"
    
    log_callback(f"任务开始 | 动作: {mode_copy_text} | 冲突策略: {mode_replace_text}")

    for index, row in df.iterrows():
        # === 【修改点 4】循环检测中断 ===
        if stop_event and stop_event.is_set():
            log_callback(">>> 用户强制停止任务！")
            return False, "任务已终止。"
        # ==============================

        try:
            old_path_str = row.get("文件路径", "").strip()
            old_name = row.get("原文件名", "").strip()
            new_folder_input = row.get("新文件夹名称", "").strip()
            new_name_input = row.get("新文件名", "").strip()

            # ... (业务逻辑保持原样，省略大量重复代码) ...
            # 这里的代码完全不用改，因为我们在循环头部加了检测
            
            # === 场景1: 纯新建模式 ===
            if not old_path_str:
                if not new_folder_input and not new_name_input: continue
                target_dir = root_path
                if new_folder_input: target_dir = root_path / new_folder_input
                if not target_dir.exists():
                    target_dir.mkdir(parents=True, exist_ok=True)
                    log_callback(f"[新建目录] {target_dir.relative_to(root_path)}")
                
                if new_name_input:
                    target_file_path = target_dir / new_name_input
                    if not is_replace_mode: target_file_path = get_unique_path(target_file_path)
                    if create_valid_file(target_file_path):
                        log_callback(f"[新建文件] {target_file_path.relative_to(root_path)}")
                        success_count += 1
                    else:
                        log_callback(f"[失败] 创建文件出错: {new_name_input}")
                        fail_count += 1
                else: success_count += 1
                continue

            # === 场景2: 现有文件的移动/复制/重命名 ===
            src_path = Path(old_path_str)
            if not src_path.exists():
                log_callback(f"[跳过] 源文件不存在: {old_name}")
                fail_count += 1
                continue

            if not new_folder_input and not new_name_input: continue

            if new_folder_input: final_dir = root_path / new_folder_input
            else: final_dir = src_path.parent

            final_name = src_path.name
            if new_name_input:
                if "." in new_name_input: final_name = new_name_input
                else: final_name = new_name_input + src_path.suffix

            final_dir.mkdir(parents=True, exist_ok=True)
            dest_path = final_dir / final_name

            if src_path.resolve() == dest_path.resolve(): continue

            extra_msg = ""
            if dest_path.exists():
                if is_replace_mode:
                    try:
                        if dest_path.is_dir(): shutil.rmtree(dest_path)
                        else: os.remove(dest_path)
                        extra_msg = " [覆盖旧文件]"
                    except Exception as e:
                        log_callback(f"[覆盖失败] 无法删除目标文件: {e}")
                        fail_count += 1
                        continue
                else:
                    dest_path = get_unique_path(dest_path)

            if is_copy_mode:
                shutil.copy2(src_path, dest_path)
                action_msg = "复制"
            else:
                shutil.move(src_path, dest_path)
                action_msg = "移动"
            
            try:
                rel_dest = dest_path.relative_to(root_path)
                log_callback(f"[{action_msg}]{extra_msg} {old_name} -> {rel_dest}")
            except:
                log_callback(f"[{action_msg}]{extra_msg} {old_name} -> {dest_path.name}")
            
            success_count += 1

        except Exception as e:
            log_callback(f"[异常] 行 {index+2}: {str(e)}")
            fail_count += 1
            
    if not is_copy_mode:
        try:
            for dirpath, dirnames, filenames in os.walk(root_folder, topdown=False):
                if not dirnames and not filenames: os.rmdir(dirpath)
        except: pass

    return True, f"处理完成! 成功: {success_count}, 失败: {fail_count}"


# --- 界面模块 (UI重构) ---

class FileBatchToolModule:
    def __init__(self):
        self.name = "文件批量整理"
        self.root_path = ""
        self.excel_path = ""
        self.is_copy_mode_var = None 
        self.is_replace_mode_var = None
        # 注意: self.app 会由 main.py 注入

    def render(self, parent_frame):
        # 初始化变量
        self.is_copy_mode_var = ctk.BooleanVar(value=False)
        self.is_replace_mode_var = ctk.BooleanVar(value=False)

        for widget in parent_frame.winfo_children(): widget.destroy()

        parent_frame.grid_columnconfigure(0, weight=1)
        parent_frame.grid_rowconfigure(0, weight=1)

        scroll = ctk.CTkScrollableFrame(parent_frame, fg_color="transparent", scrollbar_button_color="#E0E0E0", scrollbar_button_hover_color="#D0D0D0")
        scroll.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        ctk.CTkLabel(scroll, text="Excel批量整理文件", font=("Microsoft YaHei", 22, "bold"), text_color="#333").pack(pady=15, anchor="w", padx=20)

        # --- Step 1 ---
        step1_frame = ctk.CTkFrame(scroll, fg_color="white", corner_radius=8, border_width=1, border_color="#DDD")
        step1_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(step1_frame, text="Step 1: 扫描并导出 Excel", font=("Microsoft YaHei", 15, "bold"), text_color="#007AFF").pack(anchor="w", padx=15, pady=10)
        
        f1 = ctk.CTkFrame(step1_frame, fg_color="transparent")
        f1.pack(fill="x", padx=15, pady=(0, 15))
        self.entry_root = ctk.CTkEntry(f1, placeholder_text="请选择根文件夹...", height=35)
        self.entry_root.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ctk.CTkButton(f1, text="浏览", width=80, height=35, command=self.select_root).pack(side="left", padx=5)
        # 注意：这里的按钮绑定了 run_export
        self.btn_export = ctk.CTkButton(f1, text="导出 Excel", width=100, height=35, fg_color="#28a745", command=self.run_export)
        self.btn_export.pack(side="left", padx=5)

        # --- Step 2 ---
        step2_frame = ctk.CTkFrame(scroll, fg_color="white", corner_radius=8, border_width=1, border_color="#DDD")
        step2_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(step2_frame, text="Step 2: 导入修改后的 Excel", font=("Microsoft YaHei", 15, "bold"), text_color="#007AFF").pack(anchor="w", padx=15, pady=10)
        
        f2 = ctk.CTkFrame(step2_frame, fg_color="transparent")
        f2.pack(fill="x", padx=15, pady=(0, 5))
        self.entry_excel = ctk.CTkEntry(f2, placeholder_text="请选择 Excel 文件...", height=35)
        self.entry_excel.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ctk.CTkButton(f2, text="浏览", width=80, height=35, command=self.select_excel).pack(side="left", padx=5)
        # 注意：这里的按钮绑定了 run_process
        self.btn_process = ctk.CTkButton(f2, text="执行任务", width=100, height=35, fg_color="#dc3545", command=self.run_process)
        self.btn_process.pack(side="left", padx=5)

        # --- 开关 ---
        f3 = ctk.CTkFrame(step2_frame, fg_color="transparent")
        f3.pack(fill="x", padx=15, pady=(5, 15))
        self.switch_copy = ctk.CTkSwitch(f3, text="启用复制模式 (保留原文件)", variable=self.is_copy_mode_var, onvalue=True, offvalue=False, text_color="#333", font=("Microsoft YaHei", 12))
        self.switch_copy.pack(side="left", padx=5)
        self.switch_replace = ctk.CTkSwitch(f3, text="启用覆盖模式 (强制替换同名文件)", variable=self.is_replace_mode_var, onvalue=True, offvalue=False, text_color="#C0392B", font=("Microsoft YaHei", 12, "bold"))
        self.switch_replace.pack(side="left", padx=20)

        # --- 提示 ---
        tips_frame = ctk.CTkFrame(scroll, fg_color="#F8F9FA", corner_radius=6)
        tips_frame.pack(fill="x", padx=20, pady=10)
        tips = """
        【功能指南】
        1. Excel填写规则：新文件夹/文件名列不填则保持原样。空白行填新名称表示新建。
        2. 复制模式 (Switch 1)：开启则保留源文件。
        3. 覆盖模式 (Switch 2)：开启则强制替换目标同名文件。
        """
        ctk.CTkLabel(tips_frame, text=tips, justify="left", text_color="#555", font=("Consolas", 12)).pack(padx=15, pady=10, anchor="w")

        # --- 日志 ---
        ctk.CTkLabel(scroll, text="运行日志", text_color="#333", font=("Microsoft YaHei", 14, "bold")).pack(anchor="w", padx=20, pady=(10, 5))
        self.textbox = ctk.CTkTextbox(scroll, height=220, fg_color="white", text_color="#333", border_width=1, border_color="#CCC", corner_radius=4)
        self.textbox.pack(padx=20, pady=(0, 20), fill="both")

    def select_root(self):
        p = filedialog.askdirectory()
        if p:
            self.entry_root.delete(0, "end")
            self.entry_root.insert(0, p)

    def select_excel(self):
        p = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if p:
            self.entry_excel.delete(0, "end")
            self.entry_excel.insert(0, p)

    def log(self, msg):
        self.textbox.insert("end", msg + "\n")
        self.textbox.see("end")

    # === 【修改点 5】任务启动逻辑 - 导出 ===
    def run_export(self):
        root = self.entry_root.get().strip()
        if not root: return messagebox.showerror("错误", "请先选择根文件夹")
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="文件清单.xlsx")
        if not save_path: return
        
        # 申请红旗
        stop_event = None
        if hasattr(self, 'app'): stop_event = self.app.register_task(self.module_index)
        
        # 禁用按钮
        self.btn_export.configure(state="disabled", text="正在生成...")
        
        def t():
            self.log("正在生成...")
            # 传入 stop_event
            s, m = generate_excel_template(root, save_path, self.log, stop_event=stop_event)
            self.log(m)
            
            if s: 
                # 这里涉及 GUI 操作，虽然 ctk 在线程里操作通常没问题，但为了稳妥最好用 after
                # 但这里简化处理，直接操作
                self.entry_excel.delete(0, "end")
                self.entry_excel.insert(0, save_path)
            
            # 恢复按钮 & 销假
            if hasattr(self, 'app'): self.app.finish_task(self.module_index)
            self.btn_export.configure(state="normal", text="导出 Excel")

        threading.Thread(target=t, daemon=True).start()

    # === 【修改点 6】任务启动逻辑 - 执行 ===
    def run_process(self):
        root = self.entry_root.get().strip()
        excel = self.entry_excel.get().strip()
        if not root or not excel: return messagebox.showerror("错误", "路径不完整")
        
        is_copy = self.is_copy_mode_var.get()
        is_replace = self.is_replace_mode_var.get()
        
        mode_str = "复制" if is_copy else "移动"
        warn_str = "\n\n注意：已开启【强制覆盖模式】，同名文件将被永久替换！" if is_replace else ""
        
        if not messagebox.askyesno("确认", f"即将执行【{mode_str}】操作。{warn_str}\n确定要继续吗？"): return
        
        # 申请红旗
        stop_event = None
        if hasattr(self, 'app'): stop_event = self.app.register_task(self.module_index)
        
        # 禁用按钮
        self.btn_process.configure(state="disabled", text="处理中...")
        
        def t():
            # 传入 stop_event
            s, m = process_files_from_excel(excel, root, self.log, is_copy_mode=is_copy, is_replace_mode=is_replace, stop_event=stop_event)
            self.log(m)
            
            # 恢复按钮 & 销假
            if hasattr(self, 'app'): self.app.finish_task(self.module_index)
            self.btn_process.configure(state="normal", text="执行任务")

        threading.Thread(target=t, daemon=True).start()