import os
import shutil
import pandas as pd
import xlrd
import xlwt
import openpyxl
from openpyxl import Workbook, load_workbook
import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading


# --- 核心转换函数 ---

def is_html_file(file_path):
    """检测文件是否为HTML格式（伪装成xls的文件）"""
    try:
        with open(file_path, 'rb') as f:
            header = f.read(20)
            if header.startswith(b'\xef\xbb\xbf'):
                content = header[3:].decode('utf-8', errors='ignore')
            else:
                content = header.decode('utf-8', errors='ignore')
            return '<html' in content.lower() or '<!doctype html' in content.lower()
    except Exception:
        return False


def convert_html_to_xlsx_fast(src_file, dest_file):
    """
    使用Pandas的read_html进行HTML表格转换，底层C解析器，性能极高
    """
    try:
        # Pandas read_html会自动处理编码、表格解析和标签清理
        # lxml作为解析引擎，比html5lib更快
        dfs = pd.read_html(src_file, encoding='utf-8', flavor='lxml')
        
        if not dfs:
            return False, "未找到HTML表格"
        
        # 使用ExcelWriter批量写入，比openpyxl逐个cell写快得多
        with pd.ExcelWriter(dest_file, engine='openpyxl') as writer:
            for i, df in enumerate(dfs):
                # 限制sheet名长度在31字符以内
                sheet_name = f"Sheet{i+1}"
                # index=False表示不写入行号(0,1,2...)
                # header=True保留表头
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        
        return True, ""
    
    except Exception as e:
        return False, f"Pandas转换失败: {str(e)}"


def convert_xls_to_xlsx_logic(src_file, dest_file):
    """
    [模式1] xls -> xlsx (修复多Sheet丢失Bug)
    """
    # 先检测是否是 HTML 伪装的文件
    if is_html_file(src_file):
        return convert_html_to_xlsx_fast(src_file, dest_file)
    
    try:
        # 【关键修改】加上 sheet_name=None，这会读取所有 Sheet 返回一个字典
        # engine='xlrd' 用于读取老格式
        # dtype=str 防止数据类型自动转换导致精度丢失或格式错误
        all_sheets = pd.read_excel(src_file, engine='xlrd', dtype=str, sheet_name=None)
        
        # 使用 ExcelWriter 批量写入所有 Sheet
        with pd.ExcelWriter(dest_file, engine='openpyxl') as writer:
            for sheet_name, df in all_sheets.items():
                # 清理非法的 Sheet 名字符 (Excel 不允许 \ / * ? : [ ])
                safe_sheet_name = re.sub(r'[\\/*?:[\]]', '', str(sheet_name))[:31]
                df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
        
        return True, ""
    
    except Exception as e:
        # 如果 Pandas 读取失败（比如文件损坏或密码保护），尝试作为 HTML 处理
        if is_html_file(src_file):
            return convert_html_to_xlsx_fast(src_file, dest_file)
        return False, str(e)


def convert_xlsx_to_xls_logic(src_file, dest_file):
    """[模式2] xlsx -> xls"""
    try:
        wb_src = load_workbook(src_file, data_only=True)
        wb_dst = xlwt.Workbook(encoding='utf-8')
        for sheet_name in wb_src.sheetnames:
            ws_src = wb_src[sheet_name]
            ws_dst = wb_dst.add_sheet(sheet_name)
            if ws_src.max_row > 65535 or ws_src.max_column > 255:
                return False, f"超出xls限制(65k行/256列)"
            for r_idx, row in enumerate(ws_src.iter_rows(values_only=True)):
                for c_idx, value in enumerate(row):
                    if value is not None:
                        ws_dst.write(r_idx, c_idx, value)
        wb_dst.save(dest_file)
        return True, ""
    except Exception as e:
        return False, str(e)


def convert_format_change_logic(src_file, dest_file):
    """[模式3 & 4] xlsx <-> xlsm"""
    try:
        wb = load_workbook(src_file, keep_vba=True if dest_file.endswith('.xlsm') else False)
        wb.save(dest_file)
        return True, ""
    except Exception as e:
        return False, str(e)


# --- 辅助函数：文件名冲突处理 ---

def get_unique_dest_path(folder, filename):
    """如果文件已存在，生成 filename(1).ext, filename(2).ext ..."""
    base, ext = os.path.splitext(filename)
    counter = 1
    new_filename = filename
    full_path = os.path.join(folder, new_filename)
    
    while os.path.exists(full_path):
        new_filename = f"{base}({counter}){ext}"
        full_path = os.path.join(folder, new_filename)
        counter += 1
        
    return full_path


# --- 批量处理调度器 ---

def core_process_folder(src_folder, dst_folder, mode_index, copy_others, save_in_source, delete_source, log_callback, stop_event=None):
    
    if not save_in_source:
        if not os.path.exists(dst_folder):
            os.makedirs(dst_folder)
        abs_dst = os.path.abspath(dst_folder)
    else:
        abs_dst = None 

    mode_config = {
        0: (['.xls'], '.xlsx', convert_xls_to_xlsx_logic),
        1: (['.xlsx'], '.xls', convert_xlsx_to_xls_logic),
        2: (['.xlsx'], '.xlsm', convert_format_change_logic),
        3: (['.xlsm'], '.xlsx', convert_format_change_logic)
    }

    src_exts, target_ext, func = mode_config.get(mode_index)
    
    success_count = 0
    fail_count = 0
    copy_count = 0
    skipped_count = 0
    delete_count = 0
    html_converted_count = 0

    log_callback(f"开始任务: {src_exts} -> {target_ext}")
    if save_in_source:
        log_callback("输出模式: 保存在原文件同级目录")
        log_callback(f"删除原文件: {'开启' if delete_source else '关闭'}")
    else:
        log_callback(f"输出目录: {dst_folder}")
        log_callback(f"文件打包模式: {'开启' if copy_others else '关闭'}")

    for root, dirs, files in os.walk(src_folder):
        # 文件夹级中断检测
        if stop_event and stop_event.is_set():
            log_callback(">>> 用户强制停止任务！")
            return f"任务被强制终止。\n已成功: {success_count}, 失败: {fail_count}"

        if not save_in_source and abs_dst:
            abs_root = os.path.abspath(root)
            if abs_root.startswith(abs_dst):
                continue

        if save_in_source:
            target_path = root 
        else:
            relative_path = os.path.relpath(root, src_folder)
            target_path = os.path.join(dst_folder, relative_path)

        for file in files:
            # 文件级中断检测
            if stop_event and stop_event.is_set():
                log_callback(">>> 用户强制停止任务！")
                return f"任务被强制终止。\n已成功: {success_count}, 失败: {fail_count}"

            src_file_path = os.path.join(root, file)
            file_lower = file.lower()
            
            if file.startswith("~$"):
                continue

            is_target_file = any(file_lower.endswith(ext) for ext in src_exts)

            if is_target_file:
                if not os.path.exists(target_path): os.makedirs(target_path)

                base_name = os.path.splitext(file)[0]
                target_filename = base_name + target_ext
                dest_convert_path = get_unique_dest_path(target_path, target_filename)

                log_callback(f"转换: {file} -> {os.path.basename(dest_convert_path)}")
                
                status, msg = func(src_file_path, dest_convert_path)
                
                if status:
                    success_count += 1
                    if mode_index == 0 and is_html_file(src_file_path):
                        html_converted_count += 1
                        log_callback(f"  └─ [HTML伪装文件已转换]")
                    
                    if delete_source:
                        try:
                            os.remove(src_file_path)
                            delete_count += 1
                            log_callback(f"  └─ 已删除原文件: {file}")
                        except Exception as del_e:
                            log_callback(f"  └─ [删除失败] {file}: {del_e}")
                else:
                    log_callback(f"[失败] {file}: {msg}")
                    fail_count += 1
            
            else:
                if copy_others and not save_in_source:
                    if not os.path.exists(target_path): os.makedirs(target_path)
                    dest_copy_path = os.path.join(target_path, file)
                    if os.path.abspath(src_file_path) != os.path.abspath(dest_copy_path):
                        try:
                            shutil.copy2(src_file_path, dest_copy_path)
                            copy_count += 1
                        except Exception as e:
                            log_callback(f"[复制失败] {file}: {e}")
                else:
                    skipped_count += 1

    summary = f"任务结束。\n成功转换: {success_count}\n转换失败: {fail_count}"
    if html_converted_count > 0:
        summary += f"\n其中HTML伪装文件: {html_converted_count}个"
    if save_in_source:
        summary += f"\n删除原文件: {delete_count}"
    else:
        summary += f"\n打包复制: {copy_count}"
    
    return summary


# --- 界面模块 ---

class XLSToXLSXModule:
    def __init__(self):
        self.name = "Excel 格式互转"
        self.src_path = None
        self.dst_path = None
        self.modes = [
            "1. 旧版转新版 (.xls -> .xlsx)",
            "2. 新版转旧版 (.xlsx -> .xls)",
            "3. 启用宏格式 (.xlsx -> .xlsm)",
            "4. 移除宏格式 (.xlsm -> .xlsx)"
        ]

    def render(self, parent_frame):
        # 清空当前页面
        for widget in parent_frame.winfo_children():
            widget.destroy()

        # 创建全局滚动容器
        self.scroll_frame = ctk.CTkScrollableFrame(
            parent_frame,
            fg_color="transparent",
            corner_radius=0,
            scrollbar_button_color="#E0E0E0",
            scrollbar_button_hover_color="#D0D0D0"
        )
        self.scroll_frame.pack(fill="both", expand=True)

        # 标题区
        title = ctk.CTkLabel(self.scroll_frame, text="Excel格式批量互转", font=("Microsoft YaHei", 22, "bold"))
        title.pack(pady=(20, 15), anchor="w", padx=20)

        # 选项控制区 (白色卡片)
        opt_frame = ctk.CTkFrame(self.scroll_frame, fg_color="white", corner_radius=8, border_width=1, border_color="#DDD")
        opt_frame.pack(fill="x", padx=20, pady=5)

        # (1) 模式选择
        ctk.CTkLabel(opt_frame, text="转换模式:", font=("Microsoft YaHei", 14), text_color="#333").pack(anchor="w", padx=15, pady=(15, 5))
        self.combo_mode = ctk.CTkComboBox(opt_frame, values=self.modes, width=350, state="readonly")
        self.combo_mode.pack(anchor="w", padx=15, pady=(0, 10))
        self.combo_mode.set(self.modes[0])

        # (2) 复选框区
        chk_frame = ctk.CTkFrame(opt_frame, fg_color="transparent")
        chk_frame.pack(fill="x", padx=15, pady=(5, 10))

        self.var_save_in_source = ctk.BooleanVar(value=False)
        self.chk_save_source = ctk.CTkCheckBox(chk_frame, text="保存在原目录 (转换后的文件放在原文件旁边)", 
                                             variable=self.var_save_in_source, 
                                             command=self.on_save_source_change,
                                             text_color="#333", font=("Microsoft YaHei", 12))
        self.chk_save_source.pack(anchor="w", pady=4)

        self.var_delete_source = ctk.BooleanVar(value=False)
        self.chk_delete_source = ctk.CTkCheckBox(chk_frame, text="删除原文件 (转换成功后删除源文件)", 
                                               variable=self.var_delete_source, 
                                               text_color="#d63031", font=("Microsoft YaHei", 12))
        self.chk_delete_source.pack(anchor="w", pady=4)

        self.var_copy_others = ctk.BooleanVar(value=False)
        self.chk_copy = ctk.CTkCheckBox(chk_frame, text="文件打包 (复制非目标文件到输出目录)", 
                                        variable=self.var_copy_others, 
                                        text_color="#333", font=("Microsoft YaHei", 12))
        self.chk_copy.pack(anchor="w", pady=4)

        # (3) 路径选择区
        path_frame = ctk.CTkFrame(opt_frame, fg_color="transparent")
        path_frame.pack(fill="x", padx=10, pady=10)

        # 源文件夹
        f1 = ctk.CTkFrame(path_frame, fg_color="transparent")
        f1.pack(fill="x", pady=5)
        self.entry_src = ctk.CTkEntry(f1, placeholder_text="源文件夹路径", height=35)
        self.entry_src.pack(side="left", fill="x", expand=True, padx=5)
        ctk.CTkButton(f1, text="选择源", width=90, command=self.select_src).pack(side="left", padx=5)

        # 输出文件夹
        f2 = ctk.CTkFrame(path_frame, fg_color="transparent")
        f2.pack(fill="x", pady=5)
        self.entry_dst = ctk.CTkEntry(f2, placeholder_text="输出文件夹 (默认 output)", height=35)
        self.entry_dst.pack(side="left", fill="x", expand=True, padx=5)
        self.btn_dst = ctk.CTkButton(f2, text="选择输出", width=90, command=self.select_dst)
        self.btn_dst.pack(side="left", padx=5)

        # 操作按钮
        self.btn_run = ctk.CTkButton(self.scroll_frame, text="开始执行转换", command=self.run_task, 
                      fg_color="#0984e3", height=50, 
                      font=("Microsoft YaHei", 16, "bold"))
        self.btn_run.pack(pady=20, padx=20, fill="x")

        # 日志区
        ctk.CTkLabel(self.scroll_frame, text="运行日志:", anchor="w", font=("Microsoft YaHei", 12, "bold")).pack(padx=20, anchor="w")
        
        self.textbox = ctk.CTkTextbox(
            self.scroll_frame, 
            height=360,
            fg_color="white", 
            text_color="#333", 
            border_width=1, 
            border_color="#CCC",
            font=("Consolas", 12)
        )
        self.textbox.pack(padx=20, pady=(5, 30), fill="both", expand=True)

    def on_save_source_change(self):
        if self.var_save_in_source.get():
            self.entry_dst.configure(state="disabled", fg_color="#F0F0F0")
            self.btn_dst.configure(state="disabled", fg_color="#999")
            self.chk_copy.configure(state="disabled")
        else:
            self.entry_dst.configure(state="normal", fg_color="white")
            self.btn_dst.configure(state="normal", fg_color="#3B8ED0")
            self.chk_copy.configure(state="normal")

    def select_src(self):
        path = filedialog.askdirectory()
        if path:
            self.src_path = path
            self.entry_src.delete(0, "end")
            self.entry_src.insert(0, path)
            if not self.var_save_in_source.get():
                default_out = os.path.join(path, "output")
                self.entry_dst.delete(0, "end")
                self.entry_dst.insert(0, default_out)

    def select_dst(self):
        path = filedialog.askdirectory()
        if path:
            self.dst_path = path
            self.entry_dst.delete(0, "end")
            self.entry_dst.insert(0, path)

    def log(self, message):
        self.textbox.insert("end", message + "\n")
        self.textbox.see("end")

    def run_task(self):
        src = self.entry_src.get().strip()
        dst = self.entry_dst.get().strip()
        mode_str = self.combo_mode.get()
        copy_others = self.var_copy_others.get()
        save_in_source = self.var_save_in_source.get()
        delete_source = self.var_delete_source.get()
        
        mode_index = 0
        for i, m in enumerate(self.modes):
            if m == mode_str:
                mode_index = i
                break

        if not src: return messagebox.showwarning("提示", "请先选择源文件夹")
        if not save_in_source and not dst: dst = os.path.join(src, "output")

        if delete_source:
            if not messagebox.askyesno("危险操作警告", "您勾选了 [删除原文件]！\n转换成功后，原始文件将被永久删除。\n\n是否确认继续？"):
                return

        # 申请红旗
        stop_event = None
        if hasattr(self, 'app'): stop_event = self.app.register_task(self.module_index)
        
        self.btn_run.configure(state="disabled", text="转换中...")
        self.textbox.delete("1.0", "end")
        
        def thread_target():
            result = core_process_folder(
                src, dst, mode_index, copy_others, 
                save_in_source, delete_source, 
                self.log,
                stop_event=stop_event
            )
            self.log("-" * 30)
            self.log(result)
            
            # 销假 & 恢复按钮
            if hasattr(self, 'app'): self.app.finish_task(self.module_index)
            self.btn_run.configure(state="normal", text="开始执行转换")
            
            if "用户强制停止" not in result:
                messagebox.showinfo("完成", "任务执行完毕")

        threading.Thread(target=thread_target, daemon=True).start()
