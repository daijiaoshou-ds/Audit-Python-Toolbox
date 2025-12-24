import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import threading
import os
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from .core import ContraProcessor
from .algorithm import ExhaustiveSolver
from .memory import KnowledgeBase
from .occams_razor import OccamsRazor

class ContraAnalyzerUI:
    def __init__(self):
        self.name = "对方科目分析器"
        self.processor = ContraProcessor()
        self.kb = KnowledgeBase()
        self.loaded_file_path = ""
        self.map_keys = {'date': '制单日期', 'voucher_id': '凭证号', 'subject': '一级科目', 'debit': '借方金额', 'credit': '贷方金额', 'summary': '摘要'}
        self.combo_vars = {}
        self.log_box = None
        self.var_ai_pruning = None

    def render(self, parent):
        for w in parent.winfo_children(): w.destroy()
        self.main_scroll = ctk.CTkScrollableFrame(parent, fg_color="#F2F4F8", scrollbar_button_color="#E0E0E0", scrollbar_button_hover_color="#D0D0D0")
        self.main_scroll.pack(fill="both", expand=True)
        ctk.CTkLabel(self.main_scroll, text="AI 对方科目分析器 (Pro)", font=("Microsoft YaHei", 24, "bold"), text_color="#333").pack(anchor="w", padx=20, pady=(20, 10))
        self.create_load_section(self.main_scroll)
        self.create_dashboard_section(self.main_scroll)
        self.create_complex_section(self.main_scroll)
        ctk.CTkLabel(self.main_scroll, text="执行日志", font=("Arial", 12, "bold"), text_color="#555").pack(anchor="w", padx=25, pady=(10,5))
        self.log_box = ctk.CTkTextbox(self.main_scroll, height=150, fg_color="white", text_color="#333", border_color="#CCC", border_width=1, font=("Consolas", 11))
        self.log_box.pack(fill="x", padx=20, pady=(0, 30))

    def _frame(self, parent):
        f = ctk.CTkFrame(parent, fg_color="white", corner_radius=8, border_width=1, border_color="#E5E5E5")
        f.pack(fill="x", padx=20, pady=10)
        return f

    def log(self, msg):
        if self.log_box:
            self.log_box.insert("end", f"> {msg}\n")
            self.log_box.see("end")

    # ================= 1. 数据装载区 =================
    def create_load_section(self, parent):
        f = self._frame(parent)
        ctk.CTkLabel(f, text="1. 序时账导入与配置", font=("Microsoft YaHei", 15, "bold"), text_color="#007AFF").pack(anchor="w", padx=15, pady=15)
        row1 = ctk.CTkFrame(f, fg_color="transparent"); row1.pack(fill="x", padx=15)
        self.btn_load = ctk.CTkButton(row1, text="导入 Excel...", command=self.load_excel, width=120, fg_color="#F0F5FF", text_color="#007AFF", border_width=1, border_color="#007AFF"); self.btn_load.pack(side="left")
        self.lbl_file = ctk.CTkLabel(row1, text="未选择文件", text_color="#999"); self.lbl_file.pack(side="left", padx=10)
        btn_box_r = ctk.CTkFrame(row1, fg_color="transparent"); btn_box_r.pack(side="right")
        ctk.CTkButton(btn_box_r, text="清空记忆", command=self.clear_memory, fg_color="#FF9800", width=80, height=28).pack(side="left", padx=5)
        ctk.CTkButton(btn_box_r, text="重置", command=self.reset_all, fg_color="#FF4757", width=60, height=28).pack(side="left")
        self.progress_bar = ctk.CTkProgressBar(f, height=4); self.progress_bar.set(0); self.progress_bar.pack(fill="x", padx=15, pady=(15, 0))
        col_frame = ctk.CTkFrame(f, fg_color="#FAFAFA", corner_radius=6); col_frame.pack(fill="x", padx=15, pady=15)
        grid = ctk.CTkFrame(col_frame, fg_color="transparent"); grid.pack(fill="x", padx=10, pady=10); grid.grid_columnconfigure((1, 3, 5), weight=1)
        self.combo_vars = {}
        layout = [('date', 0, 0), ('voucher_id', 0, 1), ('summary', 0, 2), ('subject', 1, 0), ('debit', 1, 1), ('credit', 1, 2)]
        for key, r, c in layout:
            label = self.map_keys[key]
            ctk.CTkLabel(grid, text=f"{label}:", text_color="#333", anchor="e").grid(row=r, column=c*2, padx=5, pady=5, sticky="e")
            cb = ctk.CTkComboBox(grid, width=140, fg_color="white", button_color="#DDD", text_color="#333", dropdown_fg_color="white", dropdown_text_color="#333"); cb.set(""); cb.grid(row=r, column=c*2+1, padx=5, pady=5, sticky="ew")
            self.combo_vars[key] = cb
        self.btn_analyze = ctk.CTkButton(f, text="开始分层分析", command=self.run_analysis, height=40, font=("Microsoft YaHei", 14, "bold"), state="disabled", fg_color="#BBB"); self.btn_analyze.pack(fill="x", padx=15, pady=(0, 15))

    # ================= 2. 结果看板 =================
    def create_dashboard_section(self, parent):
        self.dash_frame = self._frame(parent)
        ctk.CTkLabel(self.dash_frame, text="2. 分析概览", font=("Microsoft YaHei", 15, "bold"), text_color="#333").pack(anchor="w", padx=15, pady=10)
        self.dash_grid = ctk.CTkFrame(self.dash_frame, fg_color="transparent"); self.dash_grid.pack(fill="x", padx=15, pady=10); self.dash_grid.grid_columnconfigure((0,1,2), weight=1)
        self.lbl_stat_total = self._make_stat_card(self.dash_grid, "总凭证数", "0", 0, "#333")
        self.lbl_stat_simple = self._make_stat_card(self.dash_grid, "自动匹配 (1v1/1vN)", "0", 1, "#00C853")
        self.lbl_stat_complex = self._make_stat_card(self.dash_grid, "复杂模式 (需穷举)", "0", 2, "#FF4757")

    def _make_stat_card(self, parent, title, val, col_idx, color):
        f = ctk.CTkFrame(parent, fg_color="#F5F7FA"); f.grid(row=0, column=col_idx, padx=5, sticky="ew")
        ctk.CTkLabel(f, text=title, text_color="gray", font=("Arial", 12)).pack(pady=(10,0))
        lbl = ctk.CTkLabel(f, text=val, font=("Arial", 22, "bold"), text_color=color); lbl.pack(pady=(0,10)); return lbl

    # ================= 3. 复杂处理中心 =================
    def create_complex_section(self, parent):
        f = self._frame(parent)
        ctk.CTkLabel(f, text="3. 复杂分录处理 (Excel 回合制)", font=("Microsoft YaHei", 15, "bold"), text_color="#007AFF").pack(anchor="w", padx=15, pady=15)
        ctk.CTkLabel(f, text="说明：系统按【合计得分】自动排序并勾选Top1。如需纠错，请在Excel中修改勾选。", text_color="#666", font=("Arial", 12)).pack(anchor="w", padx=15)
        self.complex_list_frame = ctk.CTkScrollableFrame(f, height=200, fg_color="#F9F9F9", scrollbar_button_color="#E0E0E0"); self.complex_list_frame.pack(fill="x", padx=15, pady=10)
        btn_row = ctk.CTkFrame(f, fg_color="transparent"); btn_row.pack(fill="x", padx=15, pady=15)
        self.var_ai_pruning = ctk.BooleanVar(value=True)
        self.chk_pruning = ctk.CTkCheckBox(btn_row, text="启用奥卡姆剃刀", variable=self.var_ai_pruning, text_color="#333", font=("Microsoft YaHei", 12, "bold")); self.chk_pruning.pack(side="left", padx=(0, 20))
        center_btns = ctk.CTkFrame(btn_row, fg_color="transparent"); center_btns.pack(side="left", expand=True)
        self.btn_export = ctk.CTkButton(center_btns, text="📥 导出方案到 Excel", command=self.export_all_to_excel, width=200, height=36, fg_color="#007AFF", state="disabled"); self.btn_export.pack(side="left", padx=10)
        self.btn_import = ctk.CTkButton(center_btns, text="📤 导入并生成结果", command=self.import_decisions, width=200, height=36, fg_color="#00C853", state="disabled"); self.btn_import.pack(side="left", padx=10)

    # ================= 交互逻辑 (Reset/Load/Analyze 保持不变) =================
    def reset_all(self):
        self.processor = ContraProcessor(); self.loaded_file_path = ""; self.lbl_file.configure(text="未选择"); self.log_box.delete("1.0", "end"); self.progress_bar.set(0)
        for cb in self.combo_vars.values(): cb.set("")
        self.lbl_stat_total.configure(text="0"); self.lbl_stat_simple.configure(text="0"); self.lbl_stat_complex.configure(text="0")
        for w in self.complex_list_frame.winfo_children(): w.destroy()
        self.btn_analyze.configure(state="disabled", fg_color="#BBB"); self.btn_export.configure(state="disabled"); self.btn_import.configure(state="disabled"); self.log("已重置")
    def clear_memory(self):
        if messagebox.askyesno("确认", "确定要清空记忆库吗？"): self.kb.clear_memory(); self.log("记忆库已清空。")
    def load_excel(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx;*.xls")])
        if not p: return
        self.log("正在读取表头...")
        threading.Thread(target=lambda: self.after_load(p, pd.read_excel(p, nrows=0).columns.tolist()), daemon=True).start()
    def after_load(self, path, cols):
        self.loaded_file_path = path; self.lbl_file.configure(text=os.path.basename(path)); self.log(f"文件加载成功")
        for key, cb in self.combo_vars.items():
            cb.configure(values=cols); target = self.map_keys[key]
            for c in cols:
                if target in c: cb.set(c); break
        self.btn_analyze.configure(state="normal", fg_color="#007AFF")
    def run_analysis(self):
        mapping = {}
        for k, cb in self.combo_vars.items():
            v = cb.get()
            if not v: return messagebox.showwarning("提示", f"请映射 [{self.map_keys[k]}]")
            mapping[k] = v
        self.btn_analyze.configure(state="disabled", text="分析中..."); self.progress_bar.configure(mode="indeterminate"); self.progress_bar.start()
        stop_event = None
        if hasattr(self, 'app'): stop_event = self.app.register_task(self.module_index)
        def t():
            try:
                self.log("开始数据清洗与分层..."); self.processor.load_data(self.loaded_file_path, mapping); stats = self.processor.process_all(stop_event)
                if stop_event and stop_event.is_set(): self.log("分析终止")
                else: self.update_ui_after_analysis(stats)
            except Exception as e: self.log(f"分析出错: {e}")
            finally:
                if hasattr(self, 'app'): self.app.finish_task(self.module_index)
                self.progress_bar.stop(); self.progress_bar.configure(mode="determinate"); self.progress_bar.set(1); self.btn_analyze.configure(state="normal", text="重新分析")
        threading.Thread(target=t, daemon=True).start()
    def update_ui_after_analysis(self, stats):
        self.log(f"分析完成。待人工: {stats['complex_groups']}"); self.lbl_stat_total.configure(text=str(stats['processed'])); self.lbl_stat_simple.configure(text=str(stats['simple_solved'])); self.lbl_stat_complex.configure(text=str(stats['complex_groups']))
        for w in self.complex_list_frame.winfo_children(): w.destroy()
        sorted_samples = sorted(self.processor.cluster_samples.items(), key=lambda x: x[1]['count'], reverse=True)
        for i, (k, sample) in enumerate(sorted_samples[:20]):
            row = ctk.CTkFrame(self.complex_list_frame, fg_color="white"); row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=f"Top {i+1}", width=50, text_color="gray").pack(side="left")
            ctk.CTkLabel(row, text=f"[{sample['count']}笔]", width=60, text_color="red", font=("Arial", 12, "bold")).pack(side="left")
            ctk.CTkLabel(row, text=sample['name'][:60]+"...", anchor="w", text_color="#333").pack(side="left", padx=10)
        if stats['complex_groups'] > 0: self.btn_export.configure(state="normal"); self.btn_import.configure(state="normal")
        else: self.btn_export.configure(state="disabled"); self.btn_import.configure(state="disabled")

    # ================= 核心：Excel 导出 =================
    def export_all_to_excel(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="方案选择.xlsx")
        if not path: return
        
        use_razor = self.var_ai_pruning.get()
        self.btn_export.configure(state="disabled", text="计算中...")
        self.progress_bar.configure(mode="indeterminate"); self.progress_bar.start()
        
        def t():
            try:
                solver = ExhaustiveSolver()
                all_rows = []
                total_patterns = len(self.processor.cluster_samples)
                processed = 0
                
                sorted_samples = sorted(self.processor.cluster_samples.items(), key=lambda x: x[1]['count'], reverse=True)
                
                for pattern_idx, (key_hash, sample) in enumerate(sorted_samples, 1):
                    pattern_name = sample['name']
                    
                    time.sleep(0.01)
                    solutions, is_timeout = solver.calculate_combinations(
                        sample['debits'], sample['credits'], max_solutions=200, timeout=2.0
                    )
                    
                    if not solutions: continue

                    # === 排序 ===
                    annotated_solutions = []
                    for sol in solutions:
                        r = OccamsRazor.score_solution(sol)
                        m = self.kb.get_memory_score(pattern_name, sol)
                        tot = self.kb.calculate_total_score(r, m)
                        annotated_solutions.append({"sol": sol, "razor": r, "mem": m, "total": tot})
                    
                    # 排序: Total Desc
                    annotated_solutions.sort(key=lambda x: x['total'], reverse=True)

                    # === 生成 Excel ===
                    for sol_idx, item in enumerate(annotated_solutions, 1):
                        sol = item['sol']
                        option_id = f"{pattern_idx}-{sol_idx}"
                        if is_timeout: option_id += "(超时)"
                        
                        # Top 1 自动打勾
                        check_mark = "x" if sol_idx == 1 and use_razor else ""
                        
                        desc = f"O:{item['razor']} | M:{item['mem']:.4f}"
                        if item['mem'] > 0.6: desc += " (记忆命中)"

                        all_rows.append({
                            "模式特征": pattern_name,
                            "方案ID": option_id,
                            "请在此列打x": check_mark,
                            "奥卡姆得分": item['razor'],
                            "记忆得分": item['mem'],
                            "合计得分": item['total'],
                            "会计科目": f"=== 方案 {option_id} ===",
                            "借方金额": None, "对方科目": None, "拆分金额": None, "说明": desc
                        })
                        
                        for d_subj_raw, c_map in sol.items():
                            d_name = d_subj_raw.split('__')[0]
                            valid_splits = {c: amt for c, amt in c_map.items() if abs(amt) > 0.001}
                            for c_subj_raw, amt in valid_splits.items():
                                c_name = c_subj_raw.split('__')[0]
                                all_rows.append({
                                    "模式特征": pattern_name,
                                    "方案ID": option_id,
                                    "请在此列打x": check_mark,
                                    "奥卡姆得分": None, "记忆得分": None, "合计得分": None,
                                    "会计科目": d_name,
                                    "借方金额": amt, 
                                    "对方科目": c_name,
                                    "拆分金额": amt,
                                    "说明": "明细"
                                })
                    processed += 1
                    self.progress_bar.set(processed / total_patterns)

                self.log("写入 Excel...")
                df_out = pd.DataFrame(all_rows)
                cols = ["模式特征", "方案ID", "请在此列打x", "奥卡姆得分", "记忆得分", "合计得分", "会计科目", "借方金额", "对方科目", "拆分金额", "说明"]
                for c in cols: 
                    if c not in df_out.columns: df_out[c] = ""
                df_out = df_out[cols]

                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    df_out.to_excel(writer, index=False, sheet_name="方案选择")
                    ws = writer.sheets["方案选择"]
                    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    border = Border(bottom=Side(style='thin', color="EEEEEE"))
                    font_bold = Font(bold=True, color="007AFF")
                    for row in ws.iter_rows(min_row=2):
                        if row[2].value and str(row[2].value).lower() == 'x': pass
                        if row[6].value and str(row[6].value).startswith("==="):
                            row[2].fill = fill_yellow 
                            row[2].border = border
                            row[6].font = font_bold
                    ws.column_dimensions['A'].width = 40
                    ws.column_dimensions['D'].width = 8
                    ws.column_dimensions['E'].width = 8
                    ws.column_dimensions['F'].width = 8
                    ws.column_dimensions['G'].width = 25
                    ws.column_dimensions['I'].width = 25

                self.log(f"导出成功: {path}")
                os.startfile(os.path.dirname(path))
            except Exception as e:
                self.log(f"导出错误: {e}")
                import traceback
                print(traceback.format_exc())
            finally:
                self.progress_bar.stop(); self.progress_bar.set(0)
                self.btn_export.configure(state="normal", text="📥 导出方案到 Excel")

        threading.Thread(target=t, daemon=True).start()

    # ================= 核心：导入 (内容指纹更新) =================
    def import_decisions(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not p: return
        self.log("读取规则中...")
        
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="最终对方科目分析表.xlsx")
        if not save_path: return

        self.btn_import.configure(state="disabled", text="生成最终报告...")
        self.progress_bar.configure(mode="indeterminate"); self.progress_bar.start()

        def t():
            try:
                # 1. 解析 Excel
                df = pd.read_excel(p, dtype={'方案ID': str})
                target_col = "请在此列打x"
                if target_col not in df.columns: 
                    messagebox.showerror("错误", "列名不对"); return
                
                selected_headers = df[df[target_col].notna()] 
                if selected_headers.empty:
                    self.log("警告: 未检测到任何打勾 'x'")
                    return

                learn_count = 0
                solver = ExhaustiveSolver()

                # 2. 遍历打钩的方案
                for _, row in selected_headers.iterrows():
                    pattern_name = row.get("模式特征")
                    opt_id = str(row.get("方案ID")).strip()
                    if not opt_id or opt_id.lower() == 'nan': continue

                    # === 核心：从 Excel 明细行重构【选中的方案指纹】===
                    subset = df[df["方案ID"] == opt_id]
                    # 过滤掉标题行
                    details = subset[~subset["会计科目"].astype(str).str.startswith("===")]
                    
                    if not details.empty:
                        # 从 Excel 内容重建结构: {借:{贷:1}} (金额不重要，结构重要)
                        reconstructed_sol = {}
                        for _, d_row in details.iterrows():
                            # 清洗: Excel 里显示的是不带后缀的科目名
                            # 为了生成指纹，我们直接用这些名字即可
                            # 因为 _generate_fingerprint 会自动 split('__')[0]
                            # 所以我们直接传入 "科目名" 也是兼容的
                            d = str(d_row["会计科目"]).strip()
                            c = str(d_row["对方科目"]).strip()
                            
                            # 注意: Excel 里的金额是拆分后的金额
                            # 只要有一行记录，就代表有一条边
                            if d not in reconstructed_sol: reconstructed_sol[d] = {}
                            reconstructed_sol[d][c] = 1.0 # 占位金额，用于生成指纹
                        
                        # 生成目标指纹
                        target_fingerprint = self.kb._generate_fingerprint(reconstructed_sol)
                        
                        # === 核心：获取背景板 (All Solutions) ===
                        # 为了给没选中的方案降分，我们需要重新跑一遍算法获取全量
                        # (虽然有点耗时，但这是训练过程，值得)
                        sample = None
                        for k, s in self.processor.cluster_samples.items():
                            if s['name'] == pattern_name:
                                sample = s; break
                        
                        if sample:
                            # 跑算法
                            all_solutions, _ = solver.calculate_combinations(
                                sample['debits'], sample['credits'], max_solutions=200, timeout=2.0
                            )
                            # 更新记忆 (传入指纹)
                            self.kb.update_memory_by_fingerprint(pattern_name, all_solutions, target_fingerprint)
                            learn_count += 1
                
                self.log(f"已强化记忆 {learn_count} 个模式的规则 (EMA更新)。")
                self.log("正在应用规则并生成全量数据...")
                
                # 3. 重新生成 (此时 Memory 已更新，Rank 会正确置顶)
                final_df = self.processor.finalize_report(self.kb, self.log)
                
                final_df.to_excel(save_path, index=False)
                self.log(f"最终报告生成完毕: {save_path}")
                os.startfile(os.path.dirname(save_path))
                messagebox.showinfo("完成", "所有步骤已完成！")

            except Exception as e:
                self.log(f"处理失败: {e}")
                import traceback
                print(traceback.format_exc())
            finally:
                self.progress_bar.stop(); self.progress_bar.set(0)
                self.btn_import.configure(state="normal", text="📤 导入并生成")

        threading.Thread(target=t, daemon=True).start()